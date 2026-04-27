# SPDX-License-Identifier: BSD-3-Clause
"""Template builder for igtopt – generates the gtopt Excel template from C++ headers.

The bulk of the static data (per-array field metadata, options metadata,
introduction-sheet content) lives in sibling modules:

* :mod:`igtopt._field_meta`    — ``FIELD_META`` dict keyed by array name.
* :mod:`igtopt._options_meta`  — ``*_OPTION_KEYS`` frozensets and ``_OPTIONS_FIELDS``.
* :mod:`igtopt._intro_sheet`   — ``_INTRO_LINES``, ``_EXAMPLES`` and the two
  static-content sheet builders.

Public names are re-exported from this module for back-compat with
:mod:`igtopt.igtopt` and :mod:`igtopt.tests.test_template_builder`.
"""

from __future__ import annotations

import pathlib
import re
import sys
from typing import Any

from igtopt._field_meta import FIELD_META
from igtopt._intro_sheet import (
    _EXAMPLES,
    _build_boundary_cuts_sheet,
    _build_introduction_sheet,
)
from igtopt._options_meta import (
    _OPTIONS_FIELDS,
    CASCADE_OPTION_KEYS,
    MODEL_OPTION_KEYS,
    MONOLITHIC_OPTION_KEYS,
    SDDP_OPTION_KEYS,
    SIMULATION_OPTION_KEYS,
    SOLVER_OPTION_KEYS,
)

__all__ = [
    # Re-exported data
    "FIELD_META",
    "SDDP_OPTION_KEYS",
    "CASCADE_OPTION_KEYS",
    "MONOLITHIC_OPTION_KEYS",
    "SOLVER_OPTION_KEYS",
    "MODEL_OPTION_KEYS",
    "SIMULATION_OPTION_KEYS",
    # Public functions
    "parse_json_header_fields",
    "parse_system_arrays",
    "parse_simulation_arrays",
]

# ------------------------------------------------------------------
# Ordered sheet list (determines workbook tab order)
# ------------------------------------------------------------------
_TEMPLATE_SIMULATION_SHEETS = [
    "block_array",
    "stage_array",
    "scenario_array",
    "phase_array",
    "scene_array",
    "aperture_array",
    "iteration_array",
]

_TEMPLATE_SYSTEM_SHEETS = [
    "bus_array",
    "generator_array",
    "generator_profile_array",
    "demand_array",
    "demand_profile_array",
    "line_array",
    "battery_array",
    "converter_array",
    "lng_terminal_array",
    "reserve_zone_array",
    "reserve_provision_array",
    "junction_array",
    "waterway_array",
    "flow_array",
    "reservoir_array",
    "reservoir_seepage_array",
    "reservoir_discharge_limit_array",
    "turbine_array",
    "reservoir_production_factor_array",
    "user_constraint_array",
]


def _find_repo_root(start: pathlib.Path) -> pathlib.Path:
    """Walk up the directory tree to find the repository root."""
    resolved = start.resolve()
    for parent in [resolved, *resolved.parents]:
        if (parent / "include" / "gtopt").is_dir():
            return parent
    return resolved


# Regex to extract json_member field names from JSON headers
_FIELD_NAME_RE = re.compile(r'json_\w+<"([^"]+)"')


def parse_json_header_fields(header_path: pathlib.Path) -> list[str]:
    """Extract JSON field names from a json_*.hpp header file.

    Returns the ordered list of field names as they appear in the
    json_member_list for the *last* struct defined in the file
    (which is typically the main user-facing struct, not the Attrs variant).
    """
    text = header_path.read_text(encoding="utf-8", errors="ignore")
    fields = _FIELD_NAME_RE.findall(text)
    seen: set[str] = set()
    result = []
    for f in fields:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result


def parse_system_arrays(json_system_hpp: pathlib.Path) -> list[str]:
    """Parse json_system.hpp to get the ordered list of array field names."""
    text = json_system_hpp.read_text(encoding="utf-8", errors="ignore")
    return re.findall(r'json_array_null<"([^"]+)"', text)


def parse_simulation_arrays(json_simulation_hpp: pathlib.Path) -> list[str]:
    """Parse json_simulation.hpp to get the ordered list of array field names."""
    text = json_simulation_hpp.read_text(encoding="utf-8", errors="ignore")
    return re.findall(r'json_array_null<"([^"]+)"', text)


class _WorkbookStyles:
    """Container for all Excel template styling objects."""

    def __init__(self) -> None:
        from openpyxl.styles import Font, PatternFill

        self.title_font = Font(name="Calibri", size=16, bold=True, color="1F3864")
        self.heading_font = Font(name="Calibri", size=12, bold=True, color="2F5496")
        self.body_font = Font(name="Calibri", size=10)
        self.footer_font = Font(name="Calibri", size=9, italic=True, color="808080")
        self.header_fill = PatternFill("solid", fgColor="1F3864")
        self.alt_row_fill = PatternFill("solid", fgColor="DCE6F1")
        self.help_fill = PatternFill("solid", fgColor="EBF3FB")
        self.required_fill = PatternFill("solid", fgColor="FFF2CC")
        self.header_font_white = Font(name="Calibri", bold=True, color="FFFFFF")
        self.table_header_fill = PatternFill("solid", fgColor="4472C4")
        self.table_header_font = Font(name="Calibri", bold=True, color="FFFFFF")
        self.optional_header_fill = PatternFill("solid", fgColor="2F5496")


def _build_options_sheet(wb: Any, styles: _WorkbookStyles) -> None:
    """Create and populate the options worksheet."""
    from openpyxl.styles import Alignment, Font

    ws_opts = wb.create_sheet("options")
    ws_opts.column_dimensions["A"].width = 34
    ws_opts.column_dimensions["B"].width = 22
    ws_opts.column_dimensions["C"].width = 64

    # Header row
    for col_idx, header in enumerate(["option", "value", "description"], start=1):
        c = ws_opts.cell(row=1, column=col_idx, value=header)
        c.font = styles.header_font_white
        c.fill = styles.header_fill
        c.alignment = Alignment(horizontal="left")

    for row_idx, (key, desc, default) in enumerate(_OPTIONS_FIELDS, start=2):
        ws_opts.cell(row=row_idx, column=1, value=key).font = Font(bold=False)
        if default is not None:
            ws_opts.cell(row=row_idx, column=2, value=default)
        ws_opts.cell(row=row_idx, column=3, value=desc).font = Font(
            italic=True, color="404040"
        )

    ws_opts.freeze_panes = "A2"


def _add_data_sheet(
    wb: Any,
    styles: _WorkbookStyles,
    sheet_name: str,
    fields: list[tuple[str, str, bool, str, Any]],
    examples: list[dict[str, Any]] | None = None,
) -> None:
    """Create a data worksheet with headers, help row, and example data."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws_data = wb.create_sheet(sheet_name)

    num_fields = len(fields)
    for col_idx, (fname, ftype, required, desc, _example) in enumerate(fields, start=1):
        # Header row
        header_cell = ws_data.cell(row=1, column=col_idx, value=fname)
        header_cell.font = styles.header_font_white
        if required:
            header_cell.fill = styles.header_fill  # dark blue = required
        else:
            header_cell.fill = styles.optional_header_fill  # medium
        header_cell.alignment = Alignment(horizontal="left")

        # Help row
        help_text = f"[{ftype}]{'*' if required else ''} {desc}"
        help_cell = ws_data.cell(row=2, column=col_idx, value=help_text)
        help_cell.font = Font(name="Calibri", size=8, italic=True, color="404040")
        if required:
            help_cell.fill = styles.required_fill
        else:
            help_cell.fill = styles.help_fill

    # Column widths
    for col_idx, (fname, _ft, _req, _desc, _ex) in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = max(14, len(fname) + 4)

    # Example rows
    if examples:
        for row_offset, example_row in enumerate(examples):
            data_row = 3 + row_offset
            for col_idx, (fname, _ft, _req, _desc, _default_ex) in enumerate(
                fields, start=1
            ):
                val = example_row.get(fname, None)
                if val is not None:
                    ws_data.cell(row=data_row, column=col_idx, value=val)
    else:
        # Pre-fill a single blank example row using per-field defaults
        for col_idx, (_fname, _ft, _req, _desc, example) in enumerate(fields, start=1):
            if example is not None:
                ws_data.cell(row=3, column=col_idx, value=example)

    ws_data.freeze_panes = "A3"

    # Legend in a far-right column
    legend_col = num_fields + 2
    leg_label = ws_data.cell(row=1, column=legend_col, value="Legend")
    leg_label.font = Font(bold=True)
    ws_data.cell(row=2, column=legend_col, value="Dark blue = required")
    ws_data.cell(row=3, column=legend_col, value="Medium blue = optional")
    ws_data.cell(row=4, column=legend_col, value="* = required field")
    ws_data.cell(
        row=5,
        column=legend_col,
        value="number|array|filename: scalar, inline array, or filename",
    )
    ws_data.column_dimensions[get_column_letter(legend_col)].width = 46


def _build_array_sheets(
    wb: Any,
    styles: _WorkbookStyles,
    sheet_list: list[str],
) -> None:
    """Create data sheets for a list of array names from FIELD_META."""
    for sheet in sheet_list:
        fields = FIELD_META.get(sheet, [])
        if not fields:
            continue
        examples = _EXAMPLES.get(sheet)
        _add_data_sheet(wb, styles, sheet, fields, examples)


def _build_timeseries_sheets(wb: Any, styles: _WorkbookStyles) -> None:
    """Create the example time-series worksheets (Demand@lmax, etc.)."""
    # Demand@lmax
    ws_ts = wb.create_sheet("Demand@lmax")
    for col_idx, header in enumerate(
        ["scenario", "stage", "block", "d1", "d2"], start=1
    ):
        c = ws_ts.cell(row=1, column=col_idx, value=header)
        c.font = styles.header_font_white
        c.fill = styles.header_fill
    ws_ts.cell(row=2, column=1, value=1)
    ws_ts.cell(row=2, column=2, value=1)
    ws_ts.cell(row=2, column=3, value=1)
    ws_ts.cell(row=2, column=4, value=80.0)
    ws_ts.cell(row=2, column=5, value=150.0)
    ws_ts.freeze_panes = "A2"

    # GeneratorProfile@profile
    ws_ts2 = wb.create_sheet("GeneratorProfile@profile")
    for col_idx, header in enumerate(
        ["scenario", "stage", "block", "g_solar"], start=1
    ):
        c = ws_ts2.cell(row=1, column=col_idx, value=header)
        c.font = styles.header_font_white
        c.fill = styles.header_fill
    ws_ts2.cell(row=2, column=1, value=1)
    ws_ts2.cell(row=2, column=2, value=1)
    ws_ts2.cell(row=2, column=3, value=1)
    ws_ts2.cell(row=2, column=4, value=0.75)
    ws_ts2.freeze_panes = "A2"


def _build_workbook(output_path: pathlib.Path, header_dir: pathlib.Path) -> None:
    """Build the Excel template workbook and write to *output_path*."""
    try:
        import openpyxl
    except ImportError as exc:
        print(
            f"Error: openpyxl is required.  Install it with: pip install openpyxl\n{exc}",
            file=sys.stderr,
        )
        sys.exit(1)

    styles = _WorkbookStyles()

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    default_sheet.title = ".introduction"

    _build_introduction_sheet(default_sheet, styles)
    _build_options_sheet(wb, styles)
    _build_array_sheets(wb, styles, _TEMPLATE_SIMULATION_SHEETS)
    _build_array_sheets(wb, styles, _TEMPLATE_SYSTEM_SHEETS)
    _build_timeseries_sheets(wb, styles)
    _build_boundary_cuts_sheet(wb, styles)

    wb.save(output_path)
    print(f"Template written to: {output_path}", file=sys.stderr)


def _list_sheets(header_dir: pathlib.Path) -> None:
    """Print the sheets that would be generated (does not write a file)."""
    json_dir = header_dir / "json"
    if not json_dir.is_dir():
        json_dir = header_dir

    sim_file = json_dir / "json_simulation.hpp"
    sys_file = json_dir / "json_system.hpp"

    print("# Simulation sheets (from json_simulation.hpp):")
    if sim_file.exists():
        for arr in parse_simulation_arrays(sim_file):
            print(f"  {arr}")
    else:
        for arr in _TEMPLATE_SIMULATION_SHEETS:
            print(f"  {arr}  (fallback — json_simulation.hpp not found)")

    print("\n# System sheets (from json_system.hpp):")
    if sys_file.exists():
        for arr in parse_system_arrays(sys_file):
            print(f"  {arr}")
    else:
        for arr in _TEMPLATE_SYSTEM_SHEETS:
            print(f"  {arr}  (fallback — json_system.hpp not found)")
