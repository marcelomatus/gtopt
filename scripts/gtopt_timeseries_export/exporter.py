# SPDX-License-Identifier: BSD-3-Clause
"""Excel workbook exporter for gtopt results.

The exporter produces a single ``.xlsx`` file using ``openpyxl``.  Each
output table is emitted as a separate sheet, with the sheet name derived
from the output file's parent directory and basename.  An optional
Overview sheet is populated from :mod:`gtopt_results_summary`.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pandas as pd

from gtopt_results_summary.summary import (
    _read_output_dir,
    _read_zip,
    summarize_output_dict,
)


# Maximum length allowed by Excel for a sheet name.
_MAX_SHEET_NAME = 31

# Characters not allowed in Excel sheet names.
_INVALID_CHARS = set(r":\/?*[]")


def _safe_sheet_name(name: str, seen: set[str]) -> str:
    """Normalize a sheet name to Excel constraints and avoid collisions."""
    cleaned = "".join("_" if c in _INVALID_CHARS else c for c in name)
    cleaned = cleaned.strip("_") or "sheet"
    cleaned = cleaned[:_MAX_SHEET_NAME]
    original = cleaned
    i = 2
    while cleaned.lower() in {s.lower() for s in seen}:
        suffix = f"_{i}"
        cleaned = (original[: _MAX_SHEET_NAME - len(suffix)]) + suffix
        i += 1
    seen.add(cleaned)
    return cleaned


def _autofit(worksheet) -> None:
    """Set a reasonable column width for every column in a worksheet."""
    for column_cells in worksheet.columns:
        try:
            max_len = max(
                len(str(c.value)) if c.value is not None else 0 for c in column_cells
            )
        except ValueError:
            max_len = 10
        width = min(max(10, max_len + 2), 40)
        # Column letter via the first cell
        col_letter = column_cells[0].column_letter
        worksheet.column_dimensions[col_letter].width = width


def export_from_dict(
    results: dict[str, Any],
    destination: str | os.PathLike[str],
    scale_objective: float = 1.0,
    tech_map: dict[str, str] | None = None,
) -> Path:
    """Export a parsed results dict to an Excel workbook.

    Args:
        results: Parsed results dict (shape matches
            :func:`guiservice.app._parse_results_zip`).
        destination: Output ``.xlsx`` path.
        scale_objective: Value of ``options.scale_objective`` used to unscale
            the objective value in the Overview sheet.
        tech_map: Optional per-UID technology mapping used in the Overview.

    Returns:
        The :class:`pathlib.Path` where the workbook was written.
    """
    dest = Path(destination)
    dest.parent.mkdir(parents=True, exist_ok=True)

    outputs = results.get("outputs", {})
    summary = summarize_output_dict(
        results, tech_map=tech_map, scale_objective=scale_objective
    )

    # openpyxl is an optional dependency of the scripts package (already in
    # requirements.txt).  Import here so a missing openpyxl produces a
    # clearer error message.
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for gtopt_timeseries_export; install it with "
            "`pip install openpyxl`."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1463CC")

    # Overview rows: (key, value)
    ws.append(["Metric", "Value"])
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")
    ws.append(["Status", summary.get("status")])
    ws.append(["Objective (scaled)", summary.get("obj_value")])
    ws.append(["Objective (raw)", summary.get("obj_value_raw")])
    ws.append(["Total generation (MWh)", summary.get("total_generation")])
    ws.append(["Total load (MWh)", summary.get("total_load")])
    ws.append(["Total unserved (MWh)", summary.get("total_unserved")])
    ws.append(["Peak unserved (MW)", summary.get("peak_unserved")])
    ws.append(["# Generators", summary.get("n_generators")])
    ws.append(["# Buses", summary.get("n_buses")])
    ws.append(["# Lines", summary.get("n_lines")])
    ws.append(["LMP min ($/MWh)", summary.get("lmp_min")])
    ws.append(["LMP max ($/MWh)", summary.get("lmp_max")])
    ws.append(["LMP mean ($/MWh)", summary.get("lmp_mean")])
    # Generation by tech (if any)
    tech_breakdown = summary.get("generation_by_tech", {})
    if tech_breakdown:
        ws.append([])
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value="Generation by tech (MWh)").font = Font(
            bold=True
        )
        for tech, val in tech_breakdown.items():
            ws.append([tech, val])
    _autofit(ws)

    # One sheet per output table
    seen_names: set[str] = {"Overview"}
    for key, entry in sorted(outputs.items()):
        if not isinstance(entry, dict):
            continue
        cols = entry.get("columns", [])
        data = entry.get("data", [])
        if not cols:
            continue
        sheet_name = _safe_sheet_name(key.replace("/", "_"), seen_names)
        new_ws = wb.create_sheet(title=sheet_name)
        new_ws.append(list(cols))
        for c in new_ws[1]:
            c.font = header_font
            c.fill = header_fill
        for row in data:
            new_ws.append(list(row))
        _autofit(new_ws)

    wb.save(dest)
    return dest


def export_to_excel(
    source: str | os.PathLike[str],
    destination: str | os.PathLike[str],
    scale_objective: float = 1.0,
    tech_map: dict[str, str] | None = None,
) -> Path:
    """Export a results directory or ZIP to a single Excel workbook.

    Args:
        source: Path to a directory of gtopt output files or a results ZIP.
        destination: Output ``.xlsx`` path.
        scale_objective: Value of ``options.scale_objective`` used to unscale
            the objective.
        tech_map: Optional per-UID technology mapping.

    Returns:
        The :class:`pathlib.Path` where the workbook was written.
    """
    src = Path(source)
    if not src.exists():
        raise FileNotFoundError(f"No such file or directory: {src}")
    if src.is_dir():
        results = _read_output_dir(src)
    elif src.suffix.lower() == ".zip":
        results = _read_zip(src)
    else:
        raise ValueError(
            f"Expected a directory or .zip file, got {src}"
        )
    return export_from_dict(
        results,
        destination,
        scale_objective=scale_objective,
        tech_map=tech_map,
    )
