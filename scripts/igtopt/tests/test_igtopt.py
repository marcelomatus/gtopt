"""Tests for igtopt.py – Excel → gtopt JSON conversion."""

import argparse
import csv
import json
import logging
import pathlib
import subprocess
import sys
import zipfile

import pandas as pd
import pytest

import igtopt.igtopt as _igtopt_mod  # for access to non-exported symbols
from igtopt.igtopt import (
    _run as _igtopt_run,
    create_zip_output,
    df_to_str,
    log_conversion_stats,
)

# The igtopt_c0 case (xlsx) and the reference JSON live in scripts/cases/
_SCRIPTS_DIR = pathlib.Path(__file__).parent.parent.parent
_C0_XLSX = _SCRIPTS_DIR / "cases" / "igtopt_c0" / "system_c0.xlsx"
_C0_REF_JSON = _SCRIPTS_DIR / "cases" / "json_c0" / "system_c0.json"

# The igtopt_ieee57b case (xlsx) lives in scripts/cases/igtopt_ieee57b/
_IEEE57B_XLSX = _SCRIPTS_DIR / "cases" / "igtopt_ieee57b" / "ieee57b.xlsx"

# The igtopt_bat4b24 case (xlsx) lives in scripts/cases/igtopt_bat4b24/
_BAT4B24_XLSX = _SCRIPTS_DIR / "cases" / "igtopt_bat4b24" / "bat4b24.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _run_igtopt(xlsx: pathlib.Path, tmp_path: pathlib.Path, extra_args=None):
    """Run igtopt._run() programmatically and return the parsed JSON dict."""
    json_out = tmp_path / (xlsx.stem + ".json")
    input_dir = tmp_path / xlsx.stem

    args = argparse.Namespace(
        filenames=[str(xlsx)],
        json_file=json_out,
        input_directory=input_dir,
        input_format="parquet",
        name=xlsx.stem,
        compression="gzip",
        skip_nulls=True,
        parse_unexpected_sheets=False,
        pretty=False,
        zip=False,
    )
    if extra_args:
        for k, v in extra_args.items():
            setattr(args, k, v)

    rc = _igtopt_run(args)
    assert rc == 0, "igtopt._run() returned non-zero"
    assert json_out.exists(), f"JSON output not created: {json_out}"
    return json.loads(json_out.read_text())


# ---------------------------------------------------------------------------
# Unit tests for df_to_str
# ---------------------------------------------------------------------------


def test_df_to_str_compact_format():
    """Compact format uses indent=0: produces newlines but no leading spaces."""
    df = pd.DataFrame([{"uid": 1, "name": "b1"}])
    result = df_to_str(df, skip_nulls=True)
    # With indent=0, json.dumps adds newlines but no space indentation.
    # Each line must have zero leading whitespace.
    lines = result.splitlines()
    assert all(not line.startswith(" ") for line in lines if line.strip())


def test_df_to_str_pretty_format():
    """Pretty format uses indented JSON."""
    from igtopt.igtopt import _PRETTY_INDENT, _PRETTY_SEPARATORS

    df = pd.DataFrame([{"uid": 1, "name": "b1"}])
    result = df_to_str(
        df, skip_nulls=True, indent=_PRETTY_INDENT, separators=_PRETTY_SEPARATORS
    )
    parsed = json.loads(result)
    assert parsed[0]["uid"] == 1
    assert parsed[0]["name"] == "b1"
    assert "\n" in result


def test_df_to_str_skip_nulls_drops_nan():
    """skip_nulls=True must drop NaN-valued keys from the output."""
    import numpy as np

    df = pd.DataFrame([{"uid": 1, "name": "g1", "pmax": np.nan}])
    result_skip = df_to_str(df.copy(), skip_nulls=True)
    parsed_skip = json.loads(result_skip)
    assert "pmax" not in parsed_skip[0]


def test_df_to_str_skip_dot_columns():
    """Columns whose name starts with '.' must be silently dropped."""
    df = pd.DataFrame([{"uid": 1, ".calc": "ignored"}])
    result = df_to_str(df, skip_nulls=True)
    parsed = json.loads(result)
    assert ".calc" not in parsed[0]
    assert parsed[0]["uid"] == 1


def test_df_to_str_independent_of_globals():
    """Two calls with different indent/separators must not interfere."""
    from igtopt.igtopt import _COMPACT_SEPARATORS, _PRETTY_INDENT, _PRETTY_SEPARATORS

    df = pd.DataFrame([{"uid": 1, "name": "x"}])
    pretty = df_to_str(
        df.copy(), skip_nulls=True, indent=_PRETTY_INDENT, separators=_PRETTY_SEPARATORS
    )
    compact = df_to_str(df.copy(), skip_nulls=True, separators=_COMPACT_SEPARATORS)
    # compact must be shorter (no whitespace padding)
    assert len(compact) < len(pretty)
    # both must parse to the same value
    assert json.loads(pretty) == json.loads(compact)


# ---------------------------------------------------------------------------
# Unit tests for log_conversion_stats
# ---------------------------------------------------------------------------


def test_log_conversion_stats_runs_without_error(capsys):
    """log_conversion_stats must not raise and must produce styled output."""
    counts = {
        "bus_array": 4,
        "generator_array": 3,
        "demand_array": 2,
        "line_array": 5,
        "battery_array": 1,
        "block_array": 24,
        "stage_array": 1,
        "scenario_array": 1,
    }
    opts = {
        "use_single_bus": False,
        "scale_objective": 1000,
        "demand_fail_cost": 1000,
        "input_directory": "/tmp/x",
    }
    log_conversion_stats(counts, opts, elapsed=0.5)
    captured = capsys.readouterr()
    combined = captured.out + captured.err
    assert "Buses" in combined
    assert "Generators" in combined
    assert "Elapsed" in combined


def test_log_conversion_stats_empty_counts(capsys):
    """log_conversion_stats must handle all-zero counts gracefully."""
    log_conversion_stats({}, {}, elapsed=0.0)
    captured = capsys.readouterr()
    combined = captured.out + captured.err
    assert "Buses" in combined


# ---------------------------------------------------------------------------
# Unit tests for create_zip_output
# ---------------------------------------------------------------------------


def test_create_zip_output_bundles_json_and_data(tmp_path):
    """create_zip_output must create a ZIP with the JSON + data files."""
    json_path = tmp_path / "mycase.json"
    json_path.write_text('{"name":"test"}')

    input_dir = tmp_path / "mycase"
    (input_dir / "Demand").mkdir(parents=True)
    (input_dir / "Demand" / "lmax.parquet").write_text("fake")

    zip_path = create_zip_output(json_path, input_dir)

    assert zip_path == tmp_path / "mycase.zip"
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    assert "mycase.json" in names
    assert "mycase/Demand/lmax.parquet" in names


def test_create_zip_output_no_data_dir(tmp_path):
    """create_zip_output must succeed when input_dir does not exist yet."""
    json_path = tmp_path / "empty.json"
    json_path.write_text("{}")
    input_dir = tmp_path / "nonexistent"

    zip_path = create_zip_output(json_path, input_dir)

    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        assert "empty.json" in zf.namelist()


# ---------------------------------------------------------------------------
# Unit test: _run produces a ZIP when args.zip=True
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="system_c0.xlsx not present")
def test_igtopt_c0_zip_flag_creates_archive(tmp_path):
    """Passing zip=True to _run must create a <stem>.zip archive."""
    _run_igtopt(_C0_XLSX, tmp_path, extra_args={"zip": True})
    zip_file = tmp_path / "system_c0.zip"
    assert zip_file.exists(), "ZIP archive not created when zip=True"
    with zipfile.ZipFile(zip_file) as zf:
        names = zf.namelist()
    assert "system_c0.json" in names


# ---------------------------------------------------------------------------
# Unit test: pretty formatting is independent per call
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="system_c0.xlsx not present")
def test_igtopt_pretty_compact_independent(tmp_path):
    """Two _run() calls with different pretty settings must not interfere."""
    pretty_dir = tmp_path / "pretty"
    compact_dir = tmp_path / "compact"
    pretty_dir.mkdir()
    compact_dir.mkdir()

    json_pretty = pretty_dir / "system_c0.json"
    json_compact = compact_dir / "system_c0.json"

    for json_out, input_dir, is_pretty in [
        (json_pretty, pretty_dir / "system_c0", True),
        (json_compact, compact_dir / "system_c0", False),
    ]:
        args = argparse.Namespace(
            filenames=[str(_C0_XLSX)],
            json_file=json_out,
            input_directory=input_dir,
            input_format="parquet",
            name="system_c0",
            compression="gzip",
            skip_nulls=True,
            parse_unexpected_sheets=False,
            pretty=is_pretty,
            zip=False,
        )
        rc = _igtopt_run(args)
        assert rc == 0

    # Pretty JSON should be longer (indented)
    assert json_pretty.stat().st_size > json_compact.stat().st_size
    # Both must produce valid JSON with identical bus/generator/demand arrays
    data_pretty = json.loads(json_pretty.read_text())
    data_compact = json.loads(json_compact.read_text())
    for key in ("bus_array", "generator_array", "demand_array"):
        assert data_pretty["system"].get(key) == data_compact["system"].get(key), (
            f"Mismatch in '{key}' between pretty and compact output"
        )
    for key in ("block_array",):
        assert data_pretty["simulation"].get(key) == data_compact["simulation"].get(
            key
        ), f"Mismatch in '{key}' between pretty and compact output"


# ---------------------------------------------------------------------------
# Basic smoke tests – igtopt_c0
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="system_c0.xlsx not present")
def test_igtopt_c0_produces_valid_json(tmp_path):
    data = _run_igtopt(_C0_XLSX, tmp_path)
    assert isinstance(data, dict)


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="system_c0.xlsx not present")
def test_igtopt_c0_has_options(tmp_path):
    data = _run_igtopt(_C0_XLSX, tmp_path)
    assert "options" in data
    opts = data["options"]
    assert "input_directory" in opts
    assert "input_format" in opts


# ---------------------------------------------------------------------------
# Structure matches reference json_c0
# ---------------------------------------------------------------------------


@pytest.mark.skipif(
    not _C0_XLSX.exists() or not _C0_REF_JSON.exists(),
    reason="xlsx or reference JSON not present",
)
def test_igtopt_c0_simulation_structure(tmp_path):
    """Block and stage arrays must match the reference json_c0."""
    data = _run_igtopt(_C0_XLSX, tmp_path)
    ref = json.loads(_C0_REF_JSON.read_text())

    # simulation section
    sim = data.get("simulation", data)
    ref_sim = ref.get("simulation", ref)

    for key in ("block_array", "stage_array"):
        if key in ref_sim:
            assert key in sim, f"Missing key '{key}' in igtopt output"
            assert len(sim[key]) == len(ref_sim[key]), (
                f"Length mismatch for '{key}': got {len(sim[key])}, "
                f"expected {len(ref_sim[key])}"
            )


@pytest.mark.skipif(
    not _C0_XLSX.exists() or not _C0_REF_JSON.exists(),
    reason="xlsx or reference JSON not present",
)
def test_igtopt_c0_system_structure(tmp_path):
    """Bus, generator and demand arrays must match the reference."""
    data = _run_igtopt(_C0_XLSX, tmp_path)
    ref = json.loads(_C0_REF_JSON.read_text())

    sys = data.get("system", data)
    ref_sys = ref.get("system", ref)

    for key in ("bus_array", "generator_array", "demand_array"):
        if key in ref_sys:
            assert key in sys, f"Missing key '{key}' in igtopt output"
            assert len(sys[key]) == len(ref_sys[key]), (
                f"Length mismatch for '{key}': got {len(sys[key])}, "
                f"expected {len(ref_sys[key])}"
            )


@pytest.mark.skipif(
    not _C0_XLSX.exists() or not _C0_REF_JSON.exists(),
    reason="xlsx or reference JSON not present",
)
def test_igtopt_c0_parquet_files_written(tmp_path):
    """Sheets containing '@' must produce parquet files in input_directory."""
    ref = json.loads(_C0_REF_JSON.read_text())
    # Only run if reference has an input_directory with known parquet files
    opts = ref.get("options", {})
    if "input_directory" not in opts:
        pytest.skip("no input_directory in reference options")

    _run_igtopt(_C0_XLSX, tmp_path)
    input_dir = tmp_path / "system_c0"
    # At least one .parquet file should have been created somewhere in input_dir
    parquet_files = list(input_dir.rglob("*.parquet"))
    assert len(parquet_files) > 0, "No parquet files written by igtopt"


# ---------------------------------------------------------------------------
# Integration tests – IEEE 57-bus
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_produces_valid_json(tmp_path):
    """igtopt must convert the IEEE 57-bus workbook to valid JSON."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    assert isinstance(data, dict)
    assert "system" in data
    assert "bus_array" in data["system"]
    assert "generator_array" in data["system"]
    assert "demand_array" in data["system"]
    assert "line_array" in data["system"]
    assert "options" in data


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_element_counts(tmp_path):
    """IEEE 57-bus case must have exactly 57 buses, 7 generators, 42 demands, 80 lines."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    sys = data["system"]
    assert len(sys["bus_array"]) == 57
    assert len(sys["generator_array"]) == 7
    assert len(sys["demand_array"]) == 42
    assert len(sys["line_array"]) == 80


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_simulation_structure(tmp_path):
    """IEEE 57-bus case must have 1 block, 1 stage, 1 scenario (single-snapshot OPF)."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    sim = data["simulation"]
    assert len(sim["block_array"]) == 1
    assert len(sim["stage_array"]) == 1
    assert len(sim["scenario_array"]) == 1


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_options(tmp_path):
    """IEEE 57-bus options must include model_options with use_kirchhoff etc."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    opts = data["options"]
    model = opts.get("model_options", {})
    assert model.get("use_kirchhoff") is True
    assert model.get("scale_objective") == 1000
    assert model.get("demand_fail_cost") == 1000


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_generator_fields(tmp_path):
    """All generators must have uid, name, bus, gcost, pmax, capacity fields."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    for gen in data["system"]["generator_array"]:
        for field in ("uid", "name", "bus", "gcost", "pmax", "capacity"):
            assert field in gen, f"Generator {gen.get('name', '?')} missing '{field}'"


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_line_fields(tmp_path):
    """All lines must have uid, name, bus_a, bus_b, reactance fields."""
    data = _run_igtopt(_IEEE57B_XLSX, tmp_path)
    for line in data["system"]["line_array"]:
        for field in ("uid", "name", "bus_a", "bus_b", "reactance"):
            assert field in line, f"Line {line.get('name', '?')} missing '{field}'"


# ---------------------------------------------------------------------------
# Integration tests – bat_4b_24 (4-bus, 24-block with time-series profiles)
# ---------------------------------------------------------------------------


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_produces_valid_json(tmp_path):
    """igtopt must convert the bat_4b_24 workbook to valid JSON."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    assert isinstance(data, dict)
    assert "system" in data
    sys = data["system"]
    assert "bus_array" in sys
    assert "generator_array" in sys
    assert "battery_array" in sys
    assert "generator_profile_array" in sys
    assert "options" in data


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_element_counts(tmp_path):
    """bat_4b_24 case must have 4 buses, 3 generators, 2 demands, 5 lines, 1 battery."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    sys = data["system"]
    assert len(sys["bus_array"]) == 4
    assert len(sys["generator_array"]) == 3
    assert len(sys["demand_array"]) == 2
    assert len(sys["line_array"]) == 5
    assert len(sys["battery_array"]) == 1
    assert len(sys["generator_profile_array"]) == 1


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_24_block_simulation(tmp_path):
    """bat_4b_24 simulation must have 24 blocks in 1 stage."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    sim = data["simulation"]
    assert len(sim["block_array"]) == 24
    assert len(sim["stage_array"]) == 1
    assert len(sim["scenario_array"]) == 1


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_battery_unified_definition(tmp_path):
    """The battery must use the unified definition (bus, pmax_charge, pmax_discharge)."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    bat = data["system"]["battery_array"][0]
    assert bat["name"] == "bat1"
    assert bat["bus"] == "b3"
    assert bat["pmax_charge"] == pytest.approx(60)
    assert bat["pmax_discharge"] == pytest.approx(60)
    assert bat["emax"] == pytest.approx(200)
    assert bat["input_efficiency"] == pytest.approx(0.95)
    assert bat["output_efficiency"] == pytest.approx(0.95)


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_demand_lmax_parquet(tmp_path):
    """Demand@lmax sheet must create Demand/lmax.parquet with 24 rows.

    Columns whose names match elements in demand_array (e.g. 'd3', 'd4') are
    resolved to ``uid:N`` format so the gtopt C++ binary can find them.
    """
    import pyarrow.parquet as pq  # pylint: disable=import-outside-toplevel

    _run_igtopt(_BAT4B24_XLSX, tmp_path)
    lmax_path = tmp_path / "bat4b24" / "Demand" / "lmax.parquet"
    assert lmax_path.exists(), "Demand/lmax.parquet not created"

    table = pq.read_table(str(lmax_path))
    assert table.num_rows == 24, f"Expected 24 rows, got {table.num_rows}"
    # Must have scenario, stage, block index columns + one column per demand.
    # Element names ('d3', 'd4') are resolved to uid:N format by igtopt.
    assert "scenario" in table.column_names
    assert "stage" in table.column_names
    assert "block" in table.column_names
    # d3 has uid=1, d4 has uid=2 in bat4b24's demand_array
    assert "uid:1" in table.column_names, (
        f"'uid:1' not found – name-to-uid resolution failed. Got: {table.column_names}"
    )
    assert "uid:2" in table.column_names, (
        f"'uid:2' not found – name-to-uid resolution failed. Got: {table.column_names}"
    )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_demand_lmax_profile_values(tmp_path):
    """Demand lmax parquet must contain the correct 24-hour demand profiles.

    After name-to-uid resolution, 'd3' (uid=1) → 'uid:1' and
    'd4' (uid=2) → 'uid:2'.
    """
    import pyarrow.parquet as pq  # pylint: disable=import-outside-toplevel

    _run_igtopt(_BAT4B24_XLSX, tmp_path)
    lmax_path = tmp_path / "bat4b24" / "Demand" / "lmax.parquet"
    table = pq.read_table(str(lmax_path))
    # After name-to-uid resolution: d3 (uid=1) → uid:1, d4 (uid=2) → uid:2
    d3_vals = table.column("uid:1").to_pylist()
    d4_vals = table.column("uid:2").to_pylist()

    # Reference values from bat_4b_24.json
    _D3_LMAX = [
        30,
        28,
        27,
        27,
        28,
        32,
        40,
        55,
        70,
        80,
        85,
        88,
        90,
        88,
        84,
        80,
        82,
        88,
        100,
        110,
        105,
        95,
        75,
        50,
    ]
    _D4_LMAX = [
        20,
        18,
        17,
        17,
        18,
        22,
        28,
        38,
        48,
        55,
        58,
        60,
        62,
        60,
        57,
        55,
        56,
        60,
        68,
        75,
        72,
        65,
        50,
        32,
    ]

    assert d3_vals == pytest.approx(_D3_LMAX)
    assert d4_vals == pytest.approx(_D4_LMAX)


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_generator_profile_parquet(tmp_path):
    """GeneratorProfile@profile sheet must create GeneratorProfile/profile.parquet.

    The column name 'gp_solar' (element name) is resolved to 'uid:1' via the
    generator_profile_array name-to-uid mapping.
    """
    import pyarrow.parquet as pq  # pylint: disable=import-outside-toplevel

    _run_igtopt(_BAT4B24_XLSX, tmp_path)
    profile_path = tmp_path / "bat4b24" / "GeneratorProfile" / "profile.parquet"
    assert profile_path.exists(), "GeneratorProfile/profile.parquet not created"

    table = pq.read_table(str(profile_path))
    assert table.num_rows == 24, f"Expected 24 rows, got {table.num_rows}"
    assert "block" in table.column_names
    # 'gp_solar' has uid=1 in generator_profile_array → resolved to 'uid:1'
    assert "uid:1" in table.column_names, (
        f"'uid:1' not found – name-to-uid resolution failed. Got: {table.column_names}"
    )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_solar_profile_values(tmp_path):
    """Solar profile parquet must match the reference 24-hour profile.

    After name-to-uid resolution, 'gp_solar' (uid=1) → 'uid:1'.
    """
    import pyarrow.parquet as pq  # pylint: disable=import-outside-toplevel

    _run_igtopt(_BAT4B24_XLSX, tmp_path)
    profile_path = tmp_path / "bat4b24" / "GeneratorProfile" / "profile.parquet"
    table = pq.read_table(str(profile_path))
    # gp_solar has uid=1 in generator_profile_array → resolved to uid:1
    vals = table.column("uid:1").to_pylist()

    # Reference profile from bat_4b_24.json
    _SOLAR_PROFILE = [
        0.00,
        0.00,
        0.00,
        0.00,
        0.00,
        0.00,
        0.05,
        0.15,
        0.35,
        0.55,
        0.75,
        0.90,
        1.00,
        0.95,
        0.85,
        0.70,
        0.50,
        0.30,
        0.10,
        0.02,
        0.00,
        0.00,
        0.00,
        0.00,
    ]
    assert vals == pytest.approx(_SOLAR_PROFILE, abs=1e-6)


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_demand_references_parquet_file(tmp_path):
    """Demand lmax field in the JSON must be a string (file reference), not inline values."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    for dem in data["system"]["demand_array"]:
        assert isinstance(dem["lmax"], str), (
            f"Demand '{dem['name']}' lmax should be a file reference string, "
            f"got {type(dem['lmax'])}"
        )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_generator_profile_references_parquet_file(tmp_path):
    """GeneratorProfile profile field must be a string (file reference)."""
    data = _run_igtopt(_BAT4B24_XLSX, tmp_path)
    for gp in data["system"]["generator_profile_array"]:
        assert isinstance(gp["profile"], str), (
            f"GeneratorProfile '{gp['name']}' profile should be a file reference, "
            f"got {type(gp['profile'])}"
        )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_zip_bundles_parquet_files(tmp_path):
    """ZIP output must contain the JSON and all Parquet data files."""
    _run_igtopt(_BAT4B24_XLSX, tmp_path, extra_args={"zip": True})
    zip_path = tmp_path / "bat4b24.zip"
    assert zip_path.exists(), "ZIP archive not created"

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    assert "bat4b24.json" in names
    assert any("Demand/lmax.parquet" in n for n in names)
    assert any("GeneratorProfile/profile.parquet" in n for n in names)


# ---------------------------------------------------------------------------
# E2E solver tests – igtopt → gtopt → verify results
# ---------------------------------------------------------------------------
# These tests require the gtopt binary.  They are automatically skipped when
# the binary is not found.  See conftest.py for binary discovery and CI
# artifact download logic.
# ---------------------------------------------------------------------------

_GTOPT_TIMEOUT = 60  # seconds


def _run_gtopt(
    gtopt_bin: str,
    case_dir: pathlib.Path,
    json_stem: str,
    timeout: int = _GTOPT_TIMEOUT,
) -> tuple[int, str]:
    """Run gtopt on *json_stem*.json inside *case_dir*.

    Returns (returncode, stderr).
    """
    json_file = f"{json_stem}.json"
    result = subprocess.run(
        [gtopt_bin, json_file],
        cwd=str(case_dir),
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
    )
    return result.returncode, result.stderr


def _parse_solution_csv(output_dir: pathlib.Path) -> dict[str, float | int | str]:
    """Parse solution.csv (columnar or legacy key,value) into a dict.

    For the columnar format (header: scene,phase,status,obj_value,kappa)
    returns values from the first data row.
    """
    sol = output_dir / "solution.csv"
    if not sol.exists():
        return {}
    lines = [ln.strip() for ln in sol.read_text().splitlines() if ln.strip()]
    if not lines:
        return {}
    header_fields = [f.strip() for f in lines[0].split(",")]
    result: dict[str, float | int | str] = {}
    if "obj_value" in header_fields and len(header_fields) > 2:
        if len(lines) > 1:
            vals = [v.strip() for v in lines[1].split(",")]
            for col_name, idx in zip(header_fields, range(len(header_fields))):
                if idx < len(vals):
                    raw = vals[idx]
                    try:
                        result[col_name] = int(raw)
                    except ValueError:
                        try:
                            result[col_name] = float(raw)
                        except ValueError:
                            result[col_name] = raw
    else:
        for line in lines:
            key, _, val = line.partition(",")
            key = key.strip()
            val = val.strip()
            try:
                result[key] = int(val)
            except ValueError:
                try:
                    result[key] = float(val)
                except ValueError:
                    result[key] = val
    return result


def _solution_status(output_dir: pathlib.Path) -> int:
    """Read the solver status from output/solution.csv (0 = optimal)."""
    row = _parse_solution_csv(output_dir)
    val = row.get("status", -1)
    try:
        return int(val)
    except (ValueError, TypeError):
        return -1


def _obj_value(output_dir: pathlib.Path, scale: float = 1000.0) -> float:
    """Read the objective value from output/solution.csv × scale."""
    row = _parse_solution_csv(output_dir)
    val = row.get("obj_value")
    if val is None:
        return float("nan")
    return float(val) * scale


def _prepare_case(xlsx: pathlib.Path, case_dir: pathlib.Path) -> pathlib.Path:
    """Convert xlsx → JSON in case_dir and return the JSON path."""
    json_path = case_dir / (xlsx.stem + ".json")
    input_dir = case_dir / xlsx.stem
    args = argparse.Namespace(
        filenames=[str(xlsx)],
        json_file=json_path,
        input_directory=input_dir,
        input_format="parquet",
        name=xlsx.stem,
        compression="gzip",
        skip_nulls=True,
        parse_unexpected_sheets=False,
        pretty=False,
        zip=False,
    )
    rc = _igtopt_run(args)
    assert rc == 0, "igtopt conversion failed"
    return json_path


# Reference objective values (gtopt on the original JSON, scale_objective=1000)
# obj_value in solution.csv is the physical (unscaled) cost.
_IEEE57B_OBJ_REF = 25016000.0  # $/h  (physical cost)
_BAT4B24_OBJ_REF = 44862000.0  # $/h  (physical cost)


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_gtopt_exits_zero(gtopt_bin, tmp_path):
    """gtopt must exit 0 (no crash) on the igtopt-generated ieee57b JSON."""
    case_dir = tmp_path / "ieee57b"
    case_dir.mkdir()
    _prepare_case(_IEEE57B_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "ieee57b")
    assert rc == 0, f"gtopt exited {rc}:\n{stderr}"


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_gtopt_status_optimal(gtopt_bin, tmp_path):
    """gtopt must find an optimal solution for the igtopt-generated ieee57b case."""
    case_dir = tmp_path / "ieee57b"
    case_dir.mkdir()
    _prepare_case(_IEEE57B_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "ieee57b")
    assert rc == 0, f"gtopt crashed: {stderr}"
    status = _solution_status(case_dir / "output")
    assert status == 0, f"solver status = {status} (expected 0 = optimal)"


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_gtopt_obj_matches_reference(gtopt_bin, tmp_path):
    """Objective value must be within 0.1% of the reference case ieee_57b."""
    case_dir = tmp_path / "ieee57b"
    case_dir.mkdir()
    _prepare_case(_IEEE57B_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "ieee57b")
    assert rc == 0, f"gtopt crashed: {stderr}"
    obj = _obj_value(case_dir / "output")
    assert abs(obj - _IEEE57B_OBJ_REF) / max(1.0, abs(_IEEE57B_OBJ_REF)) < 1e-3, (
        f"Objective {obj:.2f} differs from reference {_IEEE57B_OBJ_REF:.2f} by more than 0.1%"
    )


@pytest.mark.integration
@pytest.mark.skipif(not _IEEE57B_XLSX.exists(), reason="ieee57b.xlsx not present")
def test_igtopt_ieee57b_gtopt_matches_pandapower(gtopt_bin, tmp_path):
    """Generation totals and LMPs must match the pandapower DC OPF reference."""
    pytest.importorskip("pandapower")
    sys.path.insert(0, str(_SCRIPTS_DIR))
    from gtopt_compare.main import (  # pylint: disable=import-outside-toplevel
        _compare_ieee_57b as compare_fn,
    )

    case_dir = tmp_path / "ieee57b"
    case_dir.mkdir()
    _prepare_case(_IEEE57B_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "ieee57b")
    assert rc == 0, f"gtopt crashed: {stderr}"

    passed = compare_fn(case_dir / "output", tol_mw=1.0, tol_lmp=0.5)
    assert passed, "ieee57b: pandapower comparison FAILED (see printed table above)"


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_exits_zero(gtopt_bin, tmp_path):
    """gtopt must exit 0 (no crash) on the igtopt-generated bat4b24 JSON."""
    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt exited {rc}:\n{stderr}"


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_status_optimal(gtopt_bin, tmp_path):
    """gtopt must find an optimal solution for the igtopt-generated bat4b24 case."""
    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt crashed: {stderr}"
    status = _solution_status(case_dir / "output")
    assert status == 0, f"solver status = {status} (expected 0 = optimal)"


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_obj_matches_reference(gtopt_bin, tmp_path):
    """Objective value must be within 0.1% of the reference case bat_4b_24."""
    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt crashed: {stderr}"
    obj = _obj_value(case_dir / "output")
    assert abs(obj - _BAT4B24_OBJ_REF) / max(1.0, abs(_BAT4B24_OBJ_REF)) < 1e-3, (
        f"Objective {obj:.2f} differs from reference {_BAT4B24_OBJ_REF:.2f} by more than 0.1%"
    )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_no_load_shedding(gtopt_bin, tmp_path):
    """gtopt must serve all demand (fail_sol ≡ 0) for the bat4b24 case."""
    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt crashed: {stderr}"

    fail_path = case_dir / "output" / "Demand" / "fail_sol.csv"
    if not fail_path.exists():
        pytest.skip("fail_sol.csv not found in output")

    with open(fail_path, newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        uid_start = next(i for i, h in enumerate(header) if h.startswith("uid:"))
        for row in reader:
            for i in range(uid_start, len(row)):
                val = float(row[i])
                assert val < 1e-6, (
                    f"Load shedding detected: fail_sol[{i}] = {val:.4f} MW"
                )


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_curtailment_check(gtopt_bin, tmp_path):
    """gtopt_check_output curtailment check runs on bat4b24 solver output.

    The bat4b24 case has a solar generator profile (gp_solar).  After solving,
    check_renewable_curtailment should run without error and report curtailment
    data (expected to be zero or very small for this case).
    """
    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    json_path = _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt crashed: {stderr}"

    from gtopt_check_output._checks import check_renewable_curtailment  # noqa: PLC0415

    planning = json.loads(json_path.read_text(encoding="utf-8"))
    results_dir = case_dir / "output"
    findings, curtailment_data = check_renewable_curtailment(results_dir, planning)

    # Should have at least one finding (INFO or WARNING)
    assert len(findings) >= 1, "curtailment check should produce at least one finding"

    # The solar profile exists → curtailment_data should have gp_solar key
    # (may be 0.0 if no curtailment occurs)
    profile_names = {
        gp["name"]
        for gp in planning.get("system", {}).get("generator_profile_array", [])
    }
    for pname in profile_names:
        assert pname in curtailment_data, f"Missing curtailment data for {pname}"
        assert curtailment_data[pname] >= 0.0, f"Negative curtailment for {pname}"


@pytest.mark.integration
@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_igtopt_bat4b24_gtopt_matches_pandapower(gtopt_bin, tmp_path):
    """Per-block generation must match the pandapower DC OPF reference (bat_4b_24)."""
    pytest.importorskip("pandapower")
    sys.path.insert(0, str(_SCRIPTS_DIR))
    from gtopt_compare.main import (  # pylint: disable=import-outside-toplevel
        _compare_bat_4b_24 as compare_fn,
    )

    case_dir = tmp_path / "bat4b24"
    case_dir.mkdir()
    _prepare_case(_BAT4B24_XLSX, case_dir)
    rc, stderr = _run_gtopt(gtopt_bin, case_dir, "bat4b24")
    assert rc == 0, f"gtopt crashed: {stderr}"

    passed = compare_fn(case_dir / "output", tol_mw=1.0, tol_lmp=0.5)
    assert passed, "bat4b24: pandapower comparison FAILED (see printed table above)"


# ---------------------------------------------------------------------------
# split_in_columns
# ---------------------------------------------------------------------------


class TestSplitInColumns:
    """Tests for split_in_columns()."""

    def test_wraps_long_list(self):
        """A list of 10 items produces multiple lines."""
        items = [str(i) for i in range(10)]
        result = _igtopt_mod.split_in_columns(items)
        assert isinstance(result, str)
        lines = result.splitlines()
        assert len(lines) >= 2

    def test_single_item(self):
        """A single-item list renders correctly."""
        result = _igtopt_mod.split_in_columns(["one"])
        assert "one" in result

    def test_empty_list(self):
        """An empty list does not crash and returns a string."""
        result = _igtopt_mod.split_in_columns([])
        assert isinstance(result, str)


# ---------------------------------------------------------------------------
# _try_parse_json
# ---------------------------------------------------------------------------


class TestTryParseJson:
    """Tests for _try_parse_json()."""

    def test_parses_integer(self):
        assert _igtopt_mod._try_parse_json("42") == 42

    def test_parses_float(self):
        assert _igtopt_mod._try_parse_json("3.14") == pytest.approx(3.14)

    def test_parses_bool_true(self):
        assert _igtopt_mod._try_parse_json("true") is True

    def test_parses_bool_false(self):
        assert _igtopt_mod._try_parse_json("false") is False

    def test_parses_null(self):
        assert _igtopt_mod._try_parse_json("null") is None

    def test_parses_json_array(self):
        assert _igtopt_mod._try_parse_json("[1, 2, 3]") == [1, 2, 3]

    def test_returns_string_on_plain_text(self):
        result = _igtopt_mod._try_parse_json("hello world")
        assert result == "hello world"

    def test_non_string_returns_as_is(self):
        """Non-string values are returned unchanged (not parsed)."""
        assert _igtopt_mod._try_parse_json(42) == 42
        assert _igtopt_mod._try_parse_json(3.14) == pytest.approx(3.14)


# ---------------------------------------------------------------------------
# df_to_opts
# ---------------------------------------------------------------------------


class TestDfToOpts:
    """Tests for df_to_opts()."""

    def test_basic_options(self):
        """Parses option/value columns correctly."""
        df = pd.DataFrame(
            {"option": ["method", "scale_objective"], "value": ["sddp", 1000.0]}
        )
        result = _igtopt_mod.df_to_opts(df, {})
        assert result.get("method") == "sddp"
        assert result.get("scale_objective") == pytest.approx(1000.0)

    def test_numeric_option_preserved(self):
        """Numeric values are preserved as-is (not re-parsed)."""
        df = pd.DataFrame({"option": ["annual_discount_rate"], "value": [0.1]})
        result = _igtopt_mod.df_to_opts(df, {})
        assert result.get("annual_discount_rate") == pytest.approx(0.1)

    def test_existing_options_overridden_by_explicit(self):
        """Values from the explicit options dict override df values for same key."""
        df = pd.DataFrame({"option": ["input_directory"], "value": ["from_df"]})
        result = _igtopt_mod.df_to_opts(df, {"input_directory": "explicit"})
        assert result.get("input_directory") == "explicit"


# ---------------------------------------------------------------------------
# ReservoirSeepage with piecewise-linear segments
# ---------------------------------------------------------------------------


class TestReservoirSeepageSegments:
    """Tests for reservoir_seepage_array with piecewise-linear segments support."""

    def test_df_to_str_seepage_segments_string_parsed(self):
        """segments column with a JSON-encoded string is parsed to a list."""
        segments_json = (
            '[{"volume":0,"slope":0.00016132,"constant":2.18918},'
            '{"volume":500000,"slope":0.0001,"constant":4.8}]'
        )
        df = pd.DataFrame(
            [
                {
                    "uid": 1,
                    "name": "filt1",
                    "waterway": 1,
                    "reservoir": 2,
                    "segments": segments_json,
                }
            ]
        )
        result = json.loads(df_to_str(df, skip_nulls=True))
        assert len(result) == 1
        segs = result[0]["segments"]
        assert isinstance(segs, list)
        assert len(segs) == 2
        assert segs[0]["volume"] == pytest.approx(0.0)
        assert segs[0]["slope"] == pytest.approx(0.00016132)
        assert segs[0]["constant"] == pytest.approx(2.18918)
        assert segs[1]["volume"] == pytest.approx(500000.0)
        assert segs[1]["slope"] == pytest.approx(0.0001)
        assert segs[1]["constant"] == pytest.approx(4.8)

    def test_df_to_str_seepage_no_segments_omitted_when_null(self):
        """When segments is NaN/None it is omitted from JSON output."""
        import numpy as np

        df = pd.DataFrame(
            [
                {
                    "uid": 1,
                    "name": "filt1",
                    "waterway": 1,
                    "reservoir": 2,
                    "slope": 0.00005,
                    "constant": 1.2,
                    "segments": np.nan,
                }
            ]
        )
        result = json.loads(df_to_str(df, skip_nulls=True))
        assert len(result) == 1
        rec = result[0]
        assert "segments" not in rec
        assert rec["slope"] == pytest.approx(0.00005)
        assert rec["constant"] == pytest.approx(1.2)

    def test_df_to_str_seepage_slope_as_schedule_array(self):
        """slope column accepts a JSON array (per-stage schedule)."""
        df = pd.DataFrame(
            [
                {
                    "uid": 1,
                    "name": "filt1",
                    "waterway": 1,
                    "reservoir": 2,
                    "slope": "[0.0001, 0.0002]",
                    "constant": "[1.0, 1.5]",
                }
            ]
        )
        result = json.loads(df_to_str(df, skip_nulls=True))
        rec = result[0]
        assert rec["slope"] == [pytest.approx(0.0001), pytest.approx(0.0002)]
        assert rec["constant"] == [pytest.approx(1.0), pytest.approx(1.5)]

    def test_igtopt_reservoir_seepage_array_with_segments(self, tmp_path):
        """igtopt converts a reservoir_seepage_array sheet with segments to valid JSON."""
        pytest.importorskip("openpyxl")
        import openpyxl

        segments_json = (
            '[{"volume":0,"slope":0.00016132,"constant":2.18918},'
            '{"volume":500000,"slope":0.0001,"constant":4.8}]'
        )
        xlsx_path = tmp_path / "hydro_filt.xlsx"
        wb = openpyxl.Workbook()
        # options sheet (required by igtopt)
        ws_opts = wb.active
        ws_opts.title = "options"
        ws_opts.append(["option", "value"])
        ws_opts.append(["use_single_bus", True])

        # minimal simulation sheets
        ws_block = wb.create_sheet("block_array")
        ws_block.append(["uid", "duration"])
        ws_block.append([1, 1.0])

        ws_stage = wb.create_sheet("stage_array")
        ws_stage.append(["uid", "first_block", "count_block"])
        ws_stage.append([1, 1, 1])

        ws_scen = wb.create_sheet("scenario_array")
        ws_scen.append(["uid", "probability_factor"])
        ws_scen.append([1, 1.0])

        # minimal system sheets
        ws_bus = wb.create_sheet("bus_array")
        ws_bus.append(["uid", "name"])
        ws_bus.append([1, "b1"])

        ws_junc = wb.create_sheet("junction_array")
        ws_junc.append(["uid", "name"])
        ws_junc.append([1, "j1"])
        ws_junc.append([2, "j2"])

        ws_ww = wb.create_sheet("waterway_array")
        ws_ww.append(["uid", "name", "junction_a", "junction_b"])
        ws_ww.append([1, "ww1", 1, 2])

        ws_res = wb.create_sheet("reservoir_array")
        ws_res.append(["uid", "name", "junction"])
        ws_res.append([2, "embalse1", 2])

        ws_filt = wb.create_sheet("reservoir_seepage_array")
        ws_filt.append(["uid", "name", "waterway", "reservoir", "segments"])
        ws_filt.append([1, "filt1", 1, 2, segments_json])

        wb.save(xlsx_path)

        args = argparse.Namespace(
            filenames=[str(xlsx_path)],
            json_file=tmp_path / "hydro_filt.json",
            input_directory=tmp_path / "hydro_filt",
            input_format="parquet",
            name="hydro_filt",
            compression="gzip",
            skip_nulls=True,
            parse_unexpected_sheets=False,
            pretty=False,
            zip=False,
        )
        rc = _igtopt_run(args)
        assert rc == 0

        data = json.loads((tmp_path / "hydro_filt.json").read_text())
        filt_array = data["system"]["reservoir_seepage_array"]
        assert len(filt_array) == 1
        filt = filt_array[0]
        assert filt["uid"] == 1
        assert filt["name"] == "filt1"
        segs = filt["segments"]
        assert isinstance(segs, list)
        assert len(segs) == 2
        assert segs[0]["volume"] == pytest.approx(0.0)
        assert segs[0]["slope"] == pytest.approx(0.00016132)
        assert segs[0]["constant"] == pytest.approx(2.18918)
        assert segs[1]["volume"] == pytest.approx(500000.0)
        assert segs[1]["slope"] == pytest.approx(0.0001)
        assert segs[1]["constant"] == pytest.approx(4.8)


# ---------------------------------------------------------------------------
# Unit tests for name-to-uid column resolution helpers
# ---------------------------------------------------------------------------


class TestArrayNameToCname:
    """Tests for _array_name_to_cname()."""

    def test_simple_names(self):
        """Single-word arrays convert correctly."""
        from igtopt.igtopt import _array_name_to_cname

        assert _array_name_to_cname("demand_array") == "Demand"
        assert _array_name_to_cname("bus_array") == "Bus"
        assert _array_name_to_cname("generator_array") == "Generator"
        assert _array_name_to_cname("line_array") == "Line"
        assert _array_name_to_cname("battery_array") == "Battery"
        assert _array_name_to_cname("reservoir_array") == "Reservoir"
        assert _array_name_to_cname("turbine_array") == "Turbine"
        assert _array_name_to_cname("junction_array") == "Junction"
        assert _array_name_to_cname("waterway_array") == "Waterway"
        assert _array_name_to_cname("reservoir_seepage_array") == "ReservoirSeepage"
        assert _array_name_to_cname("converter_array") == "Converter"

    def test_compound_names(self):
        """Multi-word snake_case arrays convert to CamelCase correctly."""
        from igtopt.igtopt import _array_name_to_cname

        assert _array_name_to_cname("generator_profile_array") == "GeneratorProfile"
        assert _array_name_to_cname("demand_profile_array") == "DemandProfile"
        assert _array_name_to_cname("reserve_zone_array") == "ReserveZone"
        assert _array_name_to_cname("reserve_provision_array") == "ReserveProvision"
        assert (
            _array_name_to_cname("reservoir_production_factor_array")
            == "ReservoirProductionFactor"
        )
        assert (
            _array_name_to_cname("reservoir_discharge_limit_array")
            == "ReservoirDischargeLimit"
        )

    def test_no_array_suffix(self):
        """Name without _array suffix uses the whole string."""
        from igtopt.igtopt import _array_name_to_cname

        assert _array_name_to_cname("demand") == "Demand"


class TestBuildNameToUidMaps:
    """Tests for _build_name_to_uid_maps()."""

    def test_basic_mapping(self):
        """Builds name→uid:N map from a simple demand_array DataFrame."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "demand_array": pd.DataFrame(
                {"uid": [1, 2], "name": ["LAJA", "ANTUCO"], "bus": [1, 2]}
            )
        }
        maps = _build_name_to_uid_maps(xls)
        assert "Demand" in maps
        assert maps["Demand"]["LAJA"] == "uid:1"
        assert maps["Demand"]["ANTUCO"] == "uid:2"

    def test_compound_type_mapping(self):
        """Builds map for generator_profile_array → GeneratorProfile."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "generator_profile_array": pd.DataFrame(
                {"uid": [1], "name": ["gp_solar"], "generator": ["g1"]}
            )
        }
        maps = _build_name_to_uid_maps(xls)
        assert "GeneratorProfile" in maps
        assert maps["GeneratorProfile"]["gp_solar"] == "uid:1"

    def test_skips_dot_sheets(self):
        """Dot-sheets (e.g. .introduction) are ignored."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            ".introduction": pd.DataFrame({"uid": [1], "name": ["x"]}),
        }
        maps = _build_name_to_uid_maps(xls)
        assert not maps

    def test_skips_at_sheets(self):
        """@-sheets are ignored when building name maps."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "Demand@lmax": pd.DataFrame({"stage": [1], "block": [1], "LAJA": [100.0]}),
        }
        maps = _build_name_to_uid_maps(xls)
        assert not maps

    def test_skips_options_sheet(self):
        """options sheet is always skipped."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "options": pd.DataFrame({"uid": [1], "name": ["x"]}),
        }
        maps = _build_name_to_uid_maps(xls)
        assert not maps

    def test_skips_missing_uid_column(self):
        """Sheets without uid column are skipped."""
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "demand_array": pd.DataFrame({"name": ["LAJA"], "bus": [1]}),
        }
        maps = _build_name_to_uid_maps(xls)
        assert not maps

    def test_duplicate_names_omitted(self):
        """A type with duplicate element names is omitted entirely.

        When element 1 and 3 share name "DUP" but element 2 has unique name
        "UNIQUE", the entire Demand type must be omitted – not just the colliding
        pair – to prevent partial/inconsistent column resolution.
        """
        from igtopt.igtopt import _build_name_to_uid_maps

        xls = {
            "demand_array": pd.DataFrame(
                {"uid": [1, 2, 3], "name": ["DUP", "UNIQUE", "DUP"]}
            )
        }
        maps = _build_name_to_uid_maps(xls)
        # Entire Demand type must be absent – 'UNIQUE' must not leak through
        assert "Demand" not in maps


class TestResolveAtSheetColumns:
    """Tests for _resolve_at_sheet_columns()."""

    def test_renames_element_name_columns(self):
        """Element-name columns are renamed to uid:N format."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame({"stage": [1], "block": [1], "LAJA": [100.0]})
        result = _resolve_at_sheet_columns(df, {"LAJA": "uid:1"})
        assert "uid:1" in result.columns
        assert "LAJA" not in result.columns

    def test_preserves_index_columns(self):
        """scenario, stage, block columns are never renamed."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame(
            {"scenario": [1], "stage": [1], "block": [1], "LAJA": [100.0]}
        )
        result = _resolve_at_sheet_columns(df, {"scenario": "uid:0", "LAJA": "uid:1"})
        assert "scenario" in result.columns
        assert "stage" in result.columns
        assert "block" in result.columns

    def test_preserves_existing_uid_columns(self):
        """Columns already in uid:N format are left unchanged."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame({"stage": [1], "block": [1], "uid:1": [100.0]})
        result = _resolve_at_sheet_columns(df, {"uid:1": "uid:99"})
        assert "uid:1" in result.columns

    def test_returns_original_when_no_match(self):
        """Returns the original DataFrame when no columns need renaming."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame({"stage": [1], "block": [1], "uid:1": [100.0]})
        result = _resolve_at_sheet_columns(df, {"LAJA": "uid:1"})
        assert result is df  # same object, no copy made

    def test_empty_name_map(self):
        """Empty name_to_uid map returns the original DataFrame unchanged."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame({"stage": [1], "block": [1], "LAJA": [100.0]})
        result = _resolve_at_sheet_columns(df, {})
        assert result is df

    def test_multiple_columns_renamed(self):
        """Multiple element-name columns are all renamed correctly."""
        from igtopt.igtopt import _resolve_at_sheet_columns

        df = pd.DataFrame(
            {"stage": [1], "block": [1], "LAJA": [100.0], "ANTUCO": [200.0]}
        )
        result = _resolve_at_sheet_columns(df, {"LAJA": "uid:1", "ANTUCO": "uid:2"})
        assert "uid:1" in result.columns
        assert "uid:2" in result.columns
        assert "LAJA" not in result.columns
        assert "ANTUCO" not in result.columns


# ---------------------------------------------------------------------------
# igtopt roundtrip for reservoir_production_factor_array (with segments)
# ---------------------------------------------------------------------------


class TestReservoirProductionFactorRoundtrip:
    """Tests that reservoir_production_factor_array with segments survives igtopt roundtrip."""

    def _make_xlsx(self, tmp_path: pathlib.Path, segments_json: str) -> pathlib.Path:
        """Build a minimal xlsx with a reservoir_production_factor_array sheet."""
        pytest.importorskip("openpyxl")
        import openpyxl

        xlsx_path = tmp_path / "res_eff.xlsx"
        wb = openpyxl.Workbook()

        ws_opts = wb.active
        ws_opts.title = "options"
        ws_opts.append(["option", "value"])
        ws_opts.append(["use_single_bus", True])

        ws_block = wb.create_sheet("block_array")
        ws_block.append(["uid", "duration"])
        ws_block.append([1, 1.0])

        ws_stage = wb.create_sheet("stage_array")
        ws_stage.append(["uid", "first_block", "count_block"])
        ws_stage.append([1, 1, 1])

        ws_scen = wb.create_sheet("scenario_array")
        ws_scen.append(["uid", "probability_factor"])
        ws_scen.append([1, 1.0])

        ws_bus = wb.create_sheet("bus_array")
        ws_bus.append(["uid", "name"])
        ws_bus.append([1, "b1"])

        ws_gen = wb.create_sheet("generator_array")
        ws_gen.append(["uid", "name", "bus", "pmax"])
        ws_gen.append([1, "turb_gen", 1, 50.0])

        ws_junc = wb.create_sheet("junction_array")
        ws_junc.append(["uid", "name"])
        ws_junc.append([1, "j1"])
        ws_junc.append([2, "j2"])

        ws_ww = wb.create_sheet("waterway_array")
        ws_ww.append(["uid", "name", "junction_a", "junction_b"])
        ws_ww.append([1, "ww1", 1, 2])

        ws_res = wb.create_sheet("reservoir_array")
        ws_res.append(["uid", "name", "junction"])
        ws_res.append([1, "res1", 2])

        ws_turb = wb.create_sheet("turbine_array")
        ws_turb.append(["uid", "name", "waterway", "generator", "production_factor"])
        ws_turb.append([1, "t1", 1, 1, 1.5])

        ws_re = wb.create_sheet("reservoir_production_factor_array")
        ws_re.append(
            [
                "uid",
                "name",
                "turbine",
                "reservoir",
                "mean_production_factor",
                "segments",
            ]
        )
        ws_re.append([1, "re_colbun", 1, 1, 1.53, segments_json])

        wb.save(xlsx_path)
        return xlsx_path

    def _run(self, xlsx_path: pathlib.Path, tmp_path: pathlib.Path) -> dict:
        """Run igtopt on xlsx_path and return parsed JSON dict."""
        args = argparse.Namespace(
            filenames=[str(xlsx_path)],
            json_file=tmp_path / "res_eff.json",
            input_directory=tmp_path / "res_eff",
            input_format="parquet",
            name="res_eff",
            compression="gzip",
            skip_nulls=True,
            parse_unexpected_sheets=False,
            pretty=False,
            zip=False,
        )
        rc = _igtopt_run(args)
        assert rc == 0, "igtopt._run() failed"
        return json.loads((tmp_path / "res_eff.json").read_text())

    def test_reservoir_production_factor_roundtrip_single_segment(self, tmp_path):
        """reservoir_production_factor_array with one segment roundtrips correctly."""
        pytest.importorskip("openpyxl")
        segs = '[{"volume":0.0,"slope":0.0002294,"constant":1.2558}]'
        xlsx = self._make_xlsx(tmp_path, segs)
        data = self._run(xlsx, tmp_path)

        re_arr = data["system"].get("reservoir_production_factor_array", [])
        assert len(re_arr) == 1
        re = re_arr[0]
        assert re["uid"] == 1
        assert re["name"] == "re_colbun"
        assert re["mean_production_factor"] == pytest.approx(1.53)
        segs_parsed = re["segments"]
        assert isinstance(segs_parsed, list)
        assert len(segs_parsed) == 1
        assert segs_parsed[0]["volume"] == pytest.approx(0.0)
        assert segs_parsed[0]["slope"] == pytest.approx(0.0002294)
        assert segs_parsed[0]["constant"] == pytest.approx(1.2558)

    def test_reservoir_production_factor_roundtrip_multi_segment(self, tmp_path):
        """reservoir_production_factor_array with multiple segments roundtrips correctly."""
        pytest.importorskip("openpyxl")
        segs = (
            '[{"volume":0.0,"slope":0.0003,"constant":1.1},'
            '{"volume":500.0,"slope":0.0001,"constant":1.5},'
            '{"volume":1000.0,"slope":0.00005,"constant":1.8}]'
        )
        xlsx = self._make_xlsx(tmp_path, segs)
        data = self._run(xlsx, tmp_path)

        re = data["system"]["reservoir_production_factor_array"][0]
        segs_parsed = re["segments"]
        assert len(segs_parsed) == 3
        assert segs_parsed[1]["volume"] == pytest.approx(500.0)
        assert segs_parsed[2]["slope"] == pytest.approx(0.00005)

    def test_reservoir_production_factor_sddp_update_skip(self, tmp_path):
        """sddp_production_factor_update_skip field is preserved in roundtrip."""
        pytest.importorskip("openpyxl")
        import openpyxl

        xlsx_path = tmp_path / "re_skip.xlsx"
        wb = openpyxl.Workbook()
        ws_opts = wb.active
        ws_opts.title = "options"
        ws_opts.append(["option", "value"])

        for sheet_data in [
            ("block_array", [["uid", "duration"], [1, 1.0]]),
            ("stage_array", [["uid", "first_block", "count_block"], [1, 1, 1]]),
            ("scenario_array", [["uid", "probability_factor"], [1, 1.0]]),
            ("bus_array", [["uid", "name"], [1, "b1"]]),
            ("generator_array", [["uid", "name", "bus", "pmax"], [1, "g1", 1, 10.0]]),
            ("junction_array", [["uid", "name"], [1, "j1"], [2, "j2"]]),
            (
                "waterway_array",
                [["uid", "name", "junction_a", "junction_b"], [1, "ww1", 1, 2]],
            ),
            ("reservoir_array", [["uid", "name", "junction"], [1, "res1", 2]]),
            (
                "turbine_array",
                [
                    ["uid", "name", "waterway", "generator", "production_factor"],
                    [1, "t1", 1, 1, 1.2],
                ],
            ),
        ]:
            ws = wb.create_sheet(sheet_data[0])
            for row in sheet_data[1]:
                ws.append(row)

        ws_re = wb.create_sheet("reservoir_production_factor_array")
        ws_re.append(
            [
                "uid",
                "name",
                "turbine",
                "reservoir",
                "mean_production_factor",
                "sddp_production_factor_update_skip",
            ]
        )
        ws_re.append([1, "re1", 1, 1, 1.4, 5])
        wb.save(xlsx_path)

        args = argparse.Namespace(
            filenames=[str(xlsx_path)],
            json_file=tmp_path / "re_skip.json",
            input_directory=tmp_path / "re_skip",
            input_format="parquet",
            name="re_skip",
            compression="gzip",
            skip_nulls=True,
            parse_unexpected_sheets=False,
            pretty=False,
            zip=False,
        )
        rc = _igtopt_run(args)
        assert rc == 0
        data = json.loads((tmp_path / "re_skip.json").read_text())
        re = data["system"]["reservoir_production_factor_array"][0]
        assert re["sddp_production_factor_update_skip"] == 5


# ---------------------------------------------------------------------------
# Name-to-uid end-to-end via bat4b24 workbook
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not _BAT4B24_XLSX.exists(), reason="bat4b24.xlsx not present")
def test_name_to_uid_resolution_end_to_end(tmp_path):
    """Element-name columns in @-sheets are resolved to uid:N in the Parquet output.

    The bat4b24.xlsx uses element names ('d3', 'd4', 'gp_solar') as column
    headers in its @-sheets.  After igtopt processes them, the Parquet files
    must use 'uid:N' format so the gtopt C++ binary can read them.
    """
    import pyarrow.parquet as pq  # pylint: disable=import-outside-toplevel

    _run_igtopt(_BAT4B24_XLSX, tmp_path)

    # Demand: d3→uid:1, d4→uid:2
    lmax = pq.read_table(str(tmp_path / "bat4b24" / "Demand" / "lmax.parquet"))
    assert "uid:1" in lmax.column_names, "d3 not resolved to uid:1"
    assert "uid:2" in lmax.column_names, "d4 not resolved to uid:2"
    assert "d3" not in lmax.column_names, "element name 'd3' still present"
    assert "d4" not in lmax.column_names, "element name 'd4' still present"

    # GeneratorProfile: gp_solar→uid:1
    prof = pq.read_table(
        str(tmp_path / "bat4b24" / "GeneratorProfile" / "profile.parquet")
    )
    assert "uid:1" in prof.column_names, "gp_solar not resolved to uid:1"
    assert "gp_solar" not in prof.column_names, "element name 'gp_solar' still present"


# ---------------------------------------------------------------------------
# Tests for --validate and --ignore-errors flags
# ---------------------------------------------------------------------------


def _make_args(xlsx, tmp_path, **kwargs):
    """Return a minimal argparse.Namespace for _igtopt_run."""
    json_out = tmp_path / (xlsx.stem + ".json")
    input_dir = tmp_path / xlsx.stem
    defaults = {
        "filenames": [str(xlsx)],
        "json_file": json_out,
        "input_directory": input_dir,
        "input_format": "parquet",
        "name": xlsx.stem,
        "compression": "gzip",
        "skip_nulls": True,
        "parse_unexpected_sheets": False,
        "pretty": False,
        "zip": False,
        "validate": False,
        "ignore_errors": False,
    }
    defaults.update(kwargs)
    return argparse.Namespace(**defaults)


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="igtopt_c0 case not present")
def test_validate_mode_passes_on_good_workbook(tmp_path):
    """--validate on a valid workbook returns 0 and writes no output."""
    args = _make_args(_C0_XLSX, tmp_path, validate=True)
    rc = _igtopt_run(args)
    assert rc == 0, "--validate should return 0 for a valid workbook"
    assert not args.json_file.exists(), "--validate must not write any output file"


def test_validate_mode_fails_on_missing_file(tmp_path):
    """--validate on a non-existent workbook returns 1."""
    missing = tmp_path / "does_not_exist.xlsx"
    args = _make_args(missing, tmp_path, validate=True)
    rc = _igtopt_run(args)
    assert rc == 1, "--validate should return 1 when the file is missing"


def test_ignore_errors_skips_missing_file(tmp_path, caplog):
    """--ignore-errors does not abort when a file is missing; returns 0."""
    missing = tmp_path / "does_not_exist.xlsx"
    with caplog.at_level(logging.ERROR):
        args = _make_args(missing, tmp_path, ignore_errors=True)
        rc = _igtopt_run(args)
    # Should complete (no crash) even with a missing file
    assert rc == 0
    assert any("not found" in msg.lower() for msg in caplog.messages)


@pytest.mark.skipif(not _C0_XLSX.exists(), reason="igtopt_c0 case not present")
def test_unexpected_sheet_warning_message(tmp_path, caplog):
    """An unexpected sheet triggers a warning that lists expected sheet names."""
    import openpyxl  # pylint: disable=import-outside-toplevel

    # Create a workbook with one unexpected sheet
    wb_path = tmp_path / "unexpected.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "not_a_known_sheet"
    ws.append(["uid", "name"])
    ws.append([1, "foo"])
    wb.save(str(wb_path))

    with caplog.at_level(logging.WARNING):
        args = _make_args(wb_path, tmp_path)
        _igtopt_run(args)

    assert any(
        "not_a_known_sheet" in msg and "Known sheets" in msg for msg in caplog.messages
    ), (
        "Expected a warning message that names the unexpected sheet and lists known sheets"
    )
