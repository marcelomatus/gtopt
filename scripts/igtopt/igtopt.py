#!/usr/bin/env python3

import argparse
import json
import logging
import pathlib
import re
import sys
import time
import warnings
import zipfile
from typing import Any
from itertools import zip_longest

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=DeprecationWarning)

try:
    from importlib.metadata import version as _pkg_version, PackageNotFoundError

    try:
        __version__ = _pkg_version("gtopt-scripts")
    except PackageNotFoundError:
        __version__ = "dev"
except ImportError:
    __version__ = "dev"

# Sheets that belong to the ``simulation`` section of the gtopt JSON schema.
# These must match the fields declared in ``json_simulation.hpp``.
_SIMULATION_SHEETS = frozenset(
    {
        "block_array",
        "stage_array",
        "scenario_array",
        "phase_array",
        "scene_array",
    }
)

# Sheets that belong to the ``system`` section of the gtopt JSON schema.
# These must match the fields declared in ``json_system.hpp``.
_SYSTEM_SHEETS = frozenset(
    {
        "bus_array",
        "demand_array",
        "generator_array",
        "line_array",
        "generator_profile_array",
        "demand_profile_array",
        "battery_array",
        "converter_array",
        "reserve_zone_array",
        "reserve_provision_array",
        "junction_array",
        "waterway_array",
        "flow_array",
        "reservoir_array",
        "filtration_array",
        "turbine_array",
        "reservoir_efficiency_array",
    }
)

expected_sheets = ["options"] + sorted(_SIMULATION_SHEETS) + sorted(_SYSTEM_SHEETS)

_COMPACT_INDENT = 0
_COMPACT_SEPARATORS = (",", ":")

_PRETTY_INDENT = 4
_PRETTY_SEPARATORS = (", ", ": ")

# Module-level defaults kept for backwards-compatibility with external code that
# may read ``json_indent`` / ``json_separators`` directly.  New code should
# use the ``indent`` / ``separators`` parameters of :func:`df_to_str` instead,
# or rely on :func:`_run` which uses its own local copies derived from
# ``args.pretty`` so that multiple calls are fully independent.
json_indent = _COMPACT_INDENT  # pylint: disable=invalid-name
json_separators = _COMPACT_SEPARATORS  # pylint: disable=invalid-name

_LOG_LEVEL_CHOICES = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]


def split_in_columns(my_list):
    columns = 3
    result = []
    for first, second, third in zip_longest(
        my_list[0::columns], my_list[1::columns], my_list[2::columns], fillvalue=""
    ):
        result.append(f"  {first:<26}{second:<26}{third}")
    return "\n".join(result)


_DESCRIPTION = f"""\
Convert an Excel workbook to gtopt JSON input files.

Key features:
  - Sheets/columns whose name starts with "." (e.g. ".calc") are skipped
  - Sheets whose name contains "@" (e.g. "Demand@lmax") are written as
    time-series data files directly to the input directory
  - The "options" sheet populates the JSON options block
  - All other sheets become JSON arrays in the output file

Expected sheet names:
{split_in_columns(expected_sheets)}"""

_EPILOG = """
examples:
  # Basic conversion – output file and input directory derived from workbook name
  igtopt system_c0.xlsx

  # Explicit output JSON and input-data directory
  igtopt system.xlsx -j /tmp/system.json -d /tmp/system_input

  # Pretty-printed JSON, skip null values, parquet output
  igtopt system.xlsx --pretty --skip-nulls -f parquet

  # Bundle JSON + data files into a ZIP archive (ready for gtopt_guisrv/websrv)
  igtopt system.xlsx --zip

  # Convert multiple workbooks in one run
  igtopt case_a.xlsx case_b.xlsx -d /data/input

  # Show debug log messages
  igtopt system.xlsx -l DEBUG

  # Generate the igtopt Excel template from C++ JSON headers
  igtopt --make-template

  # Write template to a custom path
  igtopt --make-template -j /tmp/my_template.xlsx

  # List sheets derived from C++ headers (no file written)
  igtopt --make-template --list-sheets
"""


def df_to_file(df, input_path, cname, fname, input_format, compression):
    input_dir = pathlib.Path(input_path) / cname
    input_dir.mkdir(parents=True, exist_ok=True)
    input_file = input_dir / (fname + "." + input_format)

    types = {}
    btype = np.int32 if fname == "active" else np.float64
    for c in df.columns:
        types[c] = btype

    for c in ["scenario", "stage", "block"]:
        if c in df.columns:
            types[c] = np.int32

    df = df.astype(types)

    if input_format == "csv":
        df.to_csv(input_file, index=False)
    else:
        df.to_parquet(
            input_file,
            index=False,
            compression=compression if compression != "" else None,
        )

    return input_file


def df_to_opts(df, options):
    if "option" in df.columns and "value" in df.columns:
        opts = dict(zip(df.option, df.value))
        for key, value in options.items():
            opts[key] = value
        return opts

    logging.error(
        "'options' sheet requires both 'option' or 'value' columns, not found"
    )
    sys.exit(1)


def _try_parse_json(value):
    """Try to parse a string as JSON; return the parsed value or the original string.

    This handles Excel cells whose content is a JSON-encoded scalar, list, or
    object (e.g. ``[[55.0]]``, ``[1, 2, 3]``, ``true``).  Plain strings that
    are not valid JSON are returned unchanged.
    """
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    # Only attempt parsing if the string looks like a JSON literal
    if (
        not stripped
        or stripped[0] not in ("{", "[", "t", "f", "n", "-")
        and not (stripped[0].isdigit())
    ):
        return value
    try:
        return json.loads(stripped)
    except (json.JSONDecodeError, ValueError):
        return value


def df_to_str(
    df, skip_nulls=True, indent=_COMPACT_INDENT, separators=_COMPACT_SEPARATORS
):
    """Convert a DataFrame to a JSON string representation."""
    dropc = []
    for c in df.columns:
        if c[0] == ".":
            dropc.append(c)
    df.drop(columns=dropc, inplace=True)

    types: dict[str, Any] = {}
    if "name" in df.columns:
        types["name"] = str
    for c in ["uid", "active"]:
        if c in df.columns:
            types[c] = np.int64
    df = df.astype(types)

    if skip_nulls:
        records = []
        for _, row in df.iterrows():
            rec = {}
            for col, val in row.dropna().items():
                rec[col] = _try_parse_json(val)
            records.append(rec)
        return json.dumps(records, indent=indent, separators=separators)
    return df.to_json(
        lines=False,
        orient="records",
        date_format="epoch",
        double_precision=10,
        force_ascii=True,
        date_unit="ms",
        default_handler=None,
        indent=indent,
    )


def create_zip_output(json_path: pathlib.Path, input_dir: pathlib.Path) -> pathlib.Path:
    """Create a ZIP archive containing the JSON file and all data files.

    The archive layout mirrors what gtopt_guisrv / gtopt_websrv expect:
    ``{case_name}.json`` at the root plus all files under the input directory.
    """
    zip_path = json_path.with_suffix(".zip")
    case_name = json_path.stem
    input_dir_name = input_dir.name

    logging.info("Creating ZIP archive: %s", zip_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(json_path, arcname=f"{case_name}.json")
        if input_dir.exists():
            for data_file in sorted(input_dir.rglob("*")):
                if data_file.is_file():
                    arcname = f"{input_dir_name}/{data_file.relative_to(input_dir)}"
                    zf.write(data_file, arcname=arcname)

    logging.info(
        "ZIP archive written: %s (%d bytes)",
        zip_path,
        zip_path.stat().st_size,
    )
    return zip_path


def log_conversion_stats(
    counts: dict[str, int],
    options: dict[str, Any],
    elapsed: float,
) -> None:
    """Log conversion statistics similar to plp2gtopt."""
    logging.info("=== System statistics ===")
    logging.info("  Buses           : %d", counts.get("bus_array", 0))
    logging.info("  Generators      : %d", counts.get("generator_array", 0))
    logging.info("  Generator profs : %d", counts.get("generator_profile_array", 0))
    logging.info("  Demands         : %d", counts.get("demand_array", 0))
    logging.info("  Lines           : %d", counts.get("line_array", 0))
    logging.info("  Batteries       : %d", counts.get("batterie_array", 0))
    logging.info("  Converters      : %d", counts.get("converter_array", 0))
    logging.info("  Junctions       : %d", counts.get("junction_array", 0))
    logging.info("  Reservoirs      : %d", counts.get("reservoir_array", 0))
    logging.info("  Turbines        : %d", counts.get("turbine_array", 0))
    logging.info("=== Simulation statistics ===")
    logging.info("  Blocks          : %d", counts.get("block_array", 0))
    logging.info("  Stages          : %d", counts.get("stage_array", 0))
    logging.info("  Scenarios       : %d", counts.get("scenario_array", 0))
    logging.info("=== Key options ===")
    logging.info("  use_single_bus  : %s", options.get("use_single_bus", False))
    logging.info("  scale_objective : %s", options.get("scale_objective", 1000))
    logging.info("  demand_fail_cost: %s", options.get("demand_fail_cost", 1000))
    logging.info("  input_directory : %s", options.get("input_directory", ""))
    logging.info("=== Conversion time ===")
    logging.info("  Elapsed         : %.3fs", elapsed)


# ---------------------------------------------------------------------------
# Template generation (igtopt --make-template)
# ---------------------------------------------------------------------------

# JSON-type category strings shown in help rows
_J_INT = "integer"
_J_STR = "string"
_J_NUM = "number"
_J_BOOL = "boolean (true/false)"
_J_ID = "integer or string (uid or name)"
_J_SCHED = "number | array | filename"

# Each entry: (json_type, required, description, example_value_or_None)
_COMMON_ID_FIELDS: list[tuple[str, str, bool, str, Any]] = [
    ("uid", _J_INT, True, "Unique numeric identifier", 1),
    ("name", _J_STR, True, "Human-readable element name", "elem1"),
    ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
]

#: Per-array field metadata.
#: Keys match the JSON array names (e.g. "bus_array").
#: Values are lists of (field, json_type, required, description, example).
FIELD_META: dict[str, list[tuple[str, str, bool, str, Any]]] = {
    # ------------------------------------------------------------------
    # Simulation
    # ------------------------------------------------------------------
    "block_array": [
        ("uid", _J_INT, True, "Unique block identifier", 1),
        ("name", _J_STR, False, "Optional block label", "b1"),
        (
            "duration",
            _J_NUM,
            True,
            "Block duration in hours (e.g. 1.0 for hourly)",
            1.0,
        ),
    ],
    "stage_array": [
        ("uid", _J_INT, True, "Unique stage identifier", 1),
        ("name", _J_STR, False, "Optional stage label", "s1"),
        ("active", _J_BOOL, False, "1 = active stage (default: 1)", None),
        (
            "first_block",
            _J_INT,
            True,
            "0-based index of the first block belonging to this stage",
            0,
        ),
        ("count_block", _J_INT, True, "Number of blocks in this stage", 8760),
        (
            "discount_factor",
            _J_NUM,
            False,
            "Discount factor applied to this stage's costs (default: 1.0)",
            1.0,
        ),
    ],
    "scenario_array": [
        ("uid", _J_INT, True, "Unique scenario identifier", 1),
        ("name", _J_STR, False, "Optional scenario label", "sc1"),
        ("active", _J_BOOL, False, "1 = active scenario (default: 1)", None),
        (
            "probability_factor",
            _J_NUM,
            False,
            "Probability weight of this scenario (default: 1.0)",
            1.0,
        ),
    ],
    "phase_array": [
        ("uid", _J_INT, True, "Unique phase identifier (SDDP)", 1),
        ("name", _J_STR, False, "Optional phase label", "ph1"),
        ("active", _J_BOOL, False, "1 = active (default: 1)", None),
        (
            "first_stage",
            _J_INT,
            True,
            "0-based index of the first stage in this phase",
            0,
        ),
        ("count_stage", _J_INT, True, "Number of stages in this phase", 1),
    ],
    "scene_array": [
        ("uid", _J_INT, True, "Unique scene identifier (SDDP)", 1),
        ("name", _J_STR, False, "Optional scene label", "sc1"),
        ("active", _J_BOOL, False, "1 = active (default: 1)", None),
        (
            "first_scenario",
            _J_INT,
            True,
            "0-based index of the first scenario in this scene",
            0,
        ),
        (
            "count_scenario",
            _J_INT,
            True,
            "Number of scenarios in this scene",
            1,
        ),
    ],
    # ------------------------------------------------------------------
    # System — network
    # ------------------------------------------------------------------
    "bus_array": [
        ("uid", _J_INT, True, "Unique bus identifier", 1),
        ("name", _J_STR, True, "Bus name (used to reference this bus)", "b1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        ("voltage", _J_NUM, False, "Nominal voltage level [kV]", 220.0),
        (
            "reference_theta",
            _J_NUM,
            False,
            "Fixed voltage angle for the reference bus [rad] (default: none)",
            None,
        ),
        (
            "use_kirchhoff",
            _J_BOOL,
            False,
            "Override global Kirchhoff setting for this bus (true/false)",
            None,
        ),
    ],
    "line_array": [
        ("uid", _J_INT, True, "Unique line identifier", 1),
        ("name", _J_STR, True, "Line name", "l1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "bus_a",
            _J_ID,
            True,
            "Sending-end bus uid or name",
            "b1",
        ),
        (
            "bus_b",
            _J_ID,
            True,
            "Receiving-end bus uid or name",
            "b2",
        ),
        ("voltage", _J_SCHED, False, "Line voltage [kV]", None),
        ("resistance", _J_SCHED, False, "Series resistance [Ω] (for losses)", None),
        (
            "reactance",
            _J_SCHED,
            False,
            "Series reactance [Ω] (required for Kirchhoff DC OPF)",
            None,
        ),
        (
            "lossfactor",
            _J_SCHED,
            False,
            "Linear loss factor [p.u.] (fraction of flow lost)",
            None,
        ),
        (
            "use_line_losses",
            _J_BOOL,
            False,
            "Enable resistive loss model for this line (default: global setting)",
            None,
        ),
        (
            "loss_segments",
            _J_INT,
            False,
            "Number of piecewise-linear loss segments (>1 enables quadratic model)",
            None,
        ),
        ("tmax_ba", _J_SCHED, False, "Max power flow B→A [MW]", 500.0),
        ("tmax_ab", _J_SCHED, False, "Max power flow A→B [MW]", 500.0),
        ("tcost", _J_SCHED, False, "Transmission cost [$/MWh]", None),
        ("capacity", _J_SCHED, False, "Initial installed capacity [MW]", None),
        (
            "expcap",
            _J_SCHED,
            False,
            "Capacity per expansion module [MW/module]",
            None,
        ),
        (
            "expmod",
            _J_SCHED,
            False,
            "Maximum number of expansion modules (null = no expansion)",
            None,
        ),
        (
            "capmax",
            _J_SCHED,
            False,
            "Absolute maximum capacity after expansion [MW]",
            None,
        ),
        (
            "annual_capcost",
            _J_SCHED,
            False,
            "Annualized investment cost [$/MW-year]",
            None,
        ),
        (
            "annual_derating",
            _J_SCHED,
            False,
            "Annual capacity derating factor [p.u./year]",
            None,
        ),
    ],
    # ------------------------------------------------------------------
    # System — generation
    # ------------------------------------------------------------------
    "generator_array": [
        ("uid", _J_INT, True, "Unique generator identifier", 1),
        (
            "name",
            _J_STR,
            True,
            "Generator name (used in profile and output files)",
            "g1",
        ),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        ("bus", _J_ID, True, "Connected bus uid or name", "b1"),
        ("pmin", _J_SCHED, False, "Minimum active power output [MW]", 0.0),
        ("pmax", _J_SCHED, False, "Maximum active power output [MW]", 100.0),
        (
            "lossfactor",
            _J_SCHED,
            False,
            "Network loss factor [p.u.] (default: 0)",
            None,
        ),
        (
            "gcost",
            _J_SCHED,
            False,
            "Variable generation cost [$/MWh]",
            30.0,
        ),
        ("capacity", _J_SCHED, False, "Initial installed capacity [MW]", None),
        (
            "expcap",
            _J_SCHED,
            False,
            "Capacity per expansion module [MW/module]",
            None,
        ),
        (
            "expmod",
            _J_SCHED,
            False,
            "Maximum number of expansion modules",
            None,
        ),
        (
            "capmax",
            _J_SCHED,
            False,
            "Absolute maximum capacity after expansion [MW]",
            None,
        ),
        (
            "annual_capcost",
            _J_SCHED,
            False,
            "Annualized investment cost [$/MW-year]",
            None,
        ),
        (
            "annual_derating",
            _J_SCHED,
            False,
            "Annual capacity derating factor [p.u./year]",
            None,
        ),
    ],
    "generator_profile_array": [
        ("uid", _J_INT, True, "Unique generator-profile identifier", 1),
        ("name", _J_STR, True, "Profile name", "gp1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "generator",
            _J_ID,
            True,
            "Generator uid or name that this profile applies to",
            "g1",
        ),
        (
            "profile",
            _J_SCHED,
            True,
            "Capacity-factor values [0-1]: scalar, array, or filename "
            "(e.g. GeneratorProfile@profile sheet → 'profile')",
            "profile",
        ),
        (
            "scost",
            _J_SCHED,
            False,
            "Spilling cost when output is curtailed below profile [$/MWh]",
            None,
        ),
    ],
    # ------------------------------------------------------------------
    # System — demand
    # ------------------------------------------------------------------
    "demand_array": [
        ("uid", _J_INT, True, "Unique demand identifier", 1),
        ("name", _J_STR, True, "Demand name (used in profile and output files)", "d1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        ("bus", _J_ID, True, "Connected bus uid or name", "b1"),
        (
            "lmax",
            _J_SCHED,
            False,
            "Maximum demand (load) [MW]: scalar, array, or filename",
            100.0,
        ),
        (
            "lossfactor",
            _J_SCHED,
            False,
            "Network loss factor [p.u.] (default: 0)",
            None,
        ),
        (
            "fcost",
            _J_SCHED,
            False,
            "Failure cost for unserved load [$/MWh] (overrides global demand_fail_cost)",
            None,
        ),
        (
            "emin",
            _J_SCHED,
            False,
            "Minimum energy that must be served per stage [MWh]",
            None,
        ),
        (
            "ecost",
            _J_SCHED,
            False,
            "Energy storage cost for demand-side storage [$/MWh]",
            None,
        ),
        ("capacity", _J_SCHED, False, "Initial installed demand capacity [MW]", None),
        (
            "expcap",
            _J_SCHED,
            False,
            "Demand capacity per expansion module [MW/module]",
            None,
        ),
        ("expmod", _J_SCHED, False, "Maximum number of expansion modules", None),
        (
            "capmax",
            _J_SCHED,
            False,
            "Absolute maximum demand capacity after expansion [MW]",
            None,
        ),
        (
            "annual_capcost",
            _J_SCHED,
            False,
            "Annualized investment cost [$/MW-year]",
            None,
        ),
        (
            "annual_derating",
            _J_SCHED,
            False,
            "Annual capacity derating factor [p.u./year]",
            None,
        ),
    ],
    "demand_profile_array": [
        ("uid", _J_INT, True, "Unique demand-profile identifier", 1),
        ("name", _J_STR, True, "Profile name", "dp1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "demand",
            _J_ID,
            True,
            "Demand uid or name that this profile scales",
            "d1",
        ),
        (
            "profile",
            _J_SCHED,
            True,
            "Load-shape values [0-1]: scalar, array, or filename",
            "lmax",
        ),
        (
            "scost",
            _J_SCHED,
            False,
            "Curtailment cost [$/MWh]",
            None,
        ),
    ],
    # ------------------------------------------------------------------
    # System — storage
    # ------------------------------------------------------------------
    "battery_array": [
        ("uid", _J_INT, True, "Unique battery identifier", 1),
        ("name", _J_STR, True, "Battery name", "bat1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        ("bus", _J_ID, False, "Connected bus uid or name (optional)", None),
        (
            "input_efficiency",
            _J_SCHED,
            False,
            "Charge efficiency [p.u.] (default: 1.0)",
            0.95,
        ),
        (
            "output_efficiency",
            _J_SCHED,
            False,
            "Discharge efficiency [p.u.] (default: 1.0)",
            0.95,
        ),
        (
            "annual_loss",
            _J_SCHED,
            False,
            "Annual self-discharge loss factor [p.u./year]",
            None,
        ),
        ("emin", _J_SCHED, False, "Minimum state-of-charge [MWh]", 0.0),
        ("emax", _J_SCHED, False, "Maximum state-of-charge [MWh]", 100.0),
        (
            "ecost",
            _J_SCHED,
            False,
            "Energy storage cost (terminal value) [$/MWh]",
            None,
        ),
        (
            "eini",
            _J_NUM,
            False,
            "Initial state-of-charge [MWh] (default: emin)",
            None,
        ),
        (
            "efin",
            _J_NUM,
            False,
            "Required final state-of-charge [MWh] (default: free)",
            None,
        ),
        (
            "pmax_charge",
            _J_SCHED,
            False,
            "Maximum charging power [MW]",
            50.0,
        ),
        (
            "pmax_discharge",
            _J_SCHED,
            False,
            "Maximum discharging power [MW]",
            50.0,
        ),
        ("gcost", _J_SCHED, False, "Discharge operation cost [$/MWh]", None),
        ("capacity", _J_SCHED, False, "Initial installed capacity [MWh]", None),
        (
            "expcap",
            _J_SCHED,
            False,
            "Capacity per expansion module [MWh/module]",
            None,
        ),
        ("expmod", _J_SCHED, False, "Maximum number of expansion modules", None),
        (
            "capmax",
            _J_SCHED,
            False,
            "Absolute maximum energy capacity [MWh]",
            None,
        ),
        (
            "annual_capcost",
            _J_SCHED,
            False,
            "Annualized investment cost [$/MWh-year]",
            None,
        ),
        (
            "annual_derating",
            _J_SCHED,
            False,
            "Annual capacity derating factor [p.u./year]",
            None,
        ),
        (
            "use_state_variable",
            _J_BOOL,
            False,
            "Link SoC across planning stages/phases (true/false, default: false)",
            None,
        ),
        (
            "daily_cycle",
            _J_BOOL,
            False,
            "Reset SoC to eini at the start of each day (true/false)",
            None,
        ),
    ],
    "converter_array": [
        ("uid", _J_INT, True, "Unique converter identifier", 1),
        ("name", _J_STR, True, "Converter name", "conv1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "battery",
            _J_ID,
            True,
            "Battery uid or name linked by this converter",
            "bat1",
        ),
        (
            "generator",
            _J_ID,
            True,
            "Generator uid or name used for battery discharge",
            "g_bat1",
        ),
        (
            "demand",
            _J_ID,
            True,
            "Demand uid or name used for battery charge",
            "d_bat1",
        ),
        (
            "conversion_rate",
            _J_SCHED,
            False,
            "Energy conversion ratio [MWh/MWh] (default: 1.0)",
            None,
        ),
        ("capacity", _J_SCHED, False, "Initial converter capacity [MW]", None),
        (
            "expcap",
            _J_SCHED,
            False,
            "Capacity per expansion module [MW/module]",
            None,
        ),
        ("expmod", _J_SCHED, False, "Maximum number of expansion modules", None),
        (
            "capmax",
            _J_SCHED,
            False,
            "Absolute maximum converter capacity [MW]",
            None,
        ),
        (
            "annual_capcost",
            _J_SCHED,
            False,
            "Annualized investment cost [$/MW-year]",
            None,
        ),
        (
            "annual_derating",
            _J_SCHED,
            False,
            "Annual capacity derating factor [p.u./year]",
            None,
        ),
    ],
    # ------------------------------------------------------------------
    # System — reserves
    # ------------------------------------------------------------------
    "reserve_zone_array": [
        ("uid", _J_INT, True, "Unique reserve zone identifier", 1),
        ("name", _J_STR, True, "Reserve zone name", "rz1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "urreq",
            _J_SCHED,
            False,
            "Up-spinning reserve requirement [MW]",
            None,
        ),
        (
            "drreq",
            _J_SCHED,
            False,
            "Down-spinning reserve requirement [MW]",
            None,
        ),
        (
            "urcost",
            _J_SCHED,
            False,
            "Up-reserve failure cost [$/MW] (overrides global reserve_fail_cost)",
            None,
        ),
        (
            "drcost",
            _J_SCHED,
            False,
            "Down-reserve failure cost [$/MW]",
            None,
        ),
    ],
    "reserve_provision_array": [
        ("uid", _J_INT, True, "Unique reserve provision identifier", 1),
        ("name", _J_STR, True, "Provision name", "rp1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "generator",
            _J_ID,
            True,
            "Generator uid or name that provides the reserve",
            "g1",
        ),
        (
            "reserve_zones",
            _J_STR,
            True,
            "Comma-separated list of reserve zone names (e.g. 'rz1,rz2')",
            "rz1",
        ),
        ("urmax", _J_SCHED, False, "Maximum up-reserve contribution [MW]", None),
        ("drmax", _J_SCHED, False, "Maximum down-reserve contribution [MW]", None),
        (
            "ur_capacity_factor",
            _J_SCHED,
            False,
            "Fraction of installed capacity available for up-reserve [p.u.]",
            None,
        ),
        (
            "dr_capacity_factor",
            _J_SCHED,
            False,
            "Fraction of installed capacity available for down-reserve [p.u.]",
            None,
        ),
        (
            "ur_provision_factor",
            _J_SCHED,
            False,
            "Fraction of urmax actually provided as reserve [p.u.]",
            None,
        ),
        (
            "dr_provision_factor",
            _J_SCHED,
            False,
            "Fraction of drmax actually provided as reserve [p.u.]",
            None,
        ),
        ("urcost", _J_SCHED, False, "Up-reserve provision cost [$/MW]", None),
        ("drcost", _J_SCHED, False, "Down-reserve provision cost [$/MW]", None),
    ],
    # ------------------------------------------------------------------
    # System — hydro
    # ------------------------------------------------------------------
    "junction_array": [
        ("uid", _J_INT, True, "Unique junction identifier", 1),
        ("name", _J_STR, True, "Junction name (hydraulic node)", "j1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "drain",
            _J_BOOL,
            False,
            "If true, excess water drains freely (spills without cost)",
            None,
        ),
    ],
    "waterway_array": [
        ("uid", _J_INT, True, "Unique waterway identifier", 1),
        ("name", _J_STR, True, "Waterway name (channel between junctions)", "ww1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "junction_a",
            _J_ID,
            True,
            "Upstream junction uid or name",
            "j1",
        ),
        (
            "junction_b",
            _J_ID,
            True,
            "Downstream junction uid or name",
            "j2",
        ),
        ("capacity", _J_SCHED, False, "Maximum flow capacity [m³/s]", None),
        (
            "lossfactor",
            _J_SCHED,
            False,
            "Water loss factor [p.u.] (fraction of flow lost in transit)",
            None,
        ),
        (
            "fmin",
            _J_SCHED,
            False,
            "Minimum flow rate [m³/s] (environmental constraint)",
            None,
        ),
        ("fmax", _J_SCHED, False, "Maximum flow rate [m³/s]", None),
    ],
    "flow_array": [
        ("uid", _J_INT, True, "Unique flow identifier", 1),
        (
            "name",
            _J_STR,
            True,
            "Flow name (exogenous inflow or outflow at a junction)",
            "f1",
        ),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "direction",
            _J_INT,
            False,
            "Flow direction: 1 = inflow (positive), -1 = outflow (negative)",
            1,
        ),
        (
            "junction",
            _J_ID,
            True,
            "Junction uid or name where this flow is applied",
            "j1",
        ),
        (
            "discharge",
            _J_SCHED,
            True,
            "Flow discharge [m³/s]: scalar, array, or filename",
            10.0,
        ),
    ],
    "reservoir_array": [
        ("uid", _J_INT, True, "Unique reservoir identifier", 1),
        ("name", _J_STR, True, "Reservoir name", "res1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "junction",
            _J_ID,
            True,
            "Hydraulic junction uid or name associated with this reservoir",
            "j1",
        ),
        (
            "spillway_capacity",
            _J_NUM,
            False,
            "Maximum spillway flow capacity [m³/s]",
            None,
        ),
        (
            "spillway_cost",
            _J_NUM,
            False,
            "Cost per unit of spilled water [$/m³/s]",
            None,
        ),
        ("capacity", _J_SCHED, False, "Reservoir storage capacity [hm³]", None),
        (
            "annual_loss",
            _J_SCHED,
            False,
            "Annual evaporation/seepage loss factor [p.u./year]",
            None,
        ),
        ("emin", _J_SCHED, False, "Minimum reservoir volume [hm³]", None),
        ("emax", _J_SCHED, False, "Maximum reservoir volume [hm³]", None),
        (
            "ecost",
            _J_SCHED,
            False,
            "Terminal energy value (water value) [$/hm³]",
            None,
        ),
        ("eini", _J_NUM, False, "Initial reservoir volume [hm³]", None),
        ("efin", _J_NUM, False, "Required final reservoir volume [hm³]", None),
        (
            "fmin",
            _J_NUM,
            False,
            "Minimum turbine discharge [m³/s]",
            None,
        ),
        (
            "fmax",
            _J_NUM,
            False,
            "Maximum turbine discharge [m³/s]",
            None,
        ),
        (
            "vol_scale",
            _J_NUM,
            False,
            "Volume scaling factor [hm³/unit] (default: 1.0)",
            None,
        ),
        (
            "flow_conversion_rate",
            _J_NUM,
            False,
            "Conversion factor from flow [m³/s] to volume [hm³/block]",
            None,
        ),
        (
            "use_state_variable",
            _J_BOOL,
            False,
            "Link reservoir volume across planning stages/phases (true/false)",
            None,
        ),
        (
            "daily_cycle",
            _J_BOOL,
            False,
            "Reset reservoir to eini at the start of each day",
            None,
        ),
    ],
    "filtration_array": [
        ("uid", _J_INT, True, "Unique filtration identifier", 1),
        ("name", _J_STR, True, "Filtration name", "filt1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "waterway",
            _J_ID,
            True,
            "Waterway uid or name that loses water via filtration",
            "ww1",
        ),
        (
            "reservoir",
            _J_ID,
            True,
            "Reservoir uid or name that receives the filtered water",
            "res1",
        ),
        (
            "slope",
            _J_NUM,
            False,
            "Linear filtration coefficient [m³/s per hm³]",
            None,
        ),
        (
            "constant",
            _J_NUM,
            False,
            "Constant filtration flow [m³/s] independent of reservoir level",
            None,
        ),
        (
            "segments",
            "JSON array",
            False,
            "Piecewise-linear segments: JSON array of {volume, slope, constant} objects",
            None,
        ),
    ],
    "turbine_array": [
        ("uid", _J_INT, True, "Unique turbine identifier", 1),
        ("name", _J_STR, True, "Turbine name", "turb1"),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "waterway",
            _J_ID,
            True,
            "Waterway uid or name associated with this turbine",
            "ww1",
        ),
        (
            "generator",
            _J_ID,
            True,
            "Generator uid or name that represents this turbine's output",
            "g_hydro",
        ),
        (
            "drain",
            _J_BOOL,
            False,
            "If true, excess water bypasses the turbine (spills)",
            None,
        ),
        (
            "conversion_rate",
            _J_SCHED,
            False,
            "Water-to-power conversion [MW/(m³/s)] (productivity)",
            1.0,
        ),
        (
            "capacity",
            _J_SCHED,
            False,
            "Maximum turbine flow [m³/s]",
            None,
        ),
        (
            "main_reservoir",
            _J_ID,
            False,
            "Reservoir uid or name that is the primary head source",
            None,
        ),
    ],
    "reservoir_efficiency_array": [
        ("uid", _J_INT, True, "Unique reservoir-efficiency identifier", 1),
        (
            "name",
            _J_STR,
            True,
            "Reservoir efficiency name",
            "re1",
        ),
        ("active", _J_INT, False, "1 = active, 0 = inactive (default: 1)", None),
        (
            "turbine",
            _J_ID,
            True,
            "Turbine uid or name that this efficiency curve applies to",
            "turb1",
        ),
        (
            "reservoir",
            _J_ID,
            True,
            "Reservoir uid or name providing the hydraulic head",
            "res1",
        ),
        (
            "mean_efficiency",
            _J_NUM,
            False,
            "Mean productivity [MW/(m³/s)] averaged over operating range",
            None,
        ),
        (
            "segments",
            "JSON array",
            False,
            "Piecewise-linear efficiency segments: JSON array of "
            "{volume, slope, constant} objects",
            None,
        ),
        (
            "sddp_efficiency_update_skip",
            _J_INT,
            False,
            "SDDP iterations between efficiency updates (default: 1)",
            None,
        ),
    ],
}

# ------------------------------------------------------------------
# Options sheet metadata
# (field, description, default_value_or_None)
# ------------------------------------------------------------------
_OPTIONS_FIELDS: list[tuple[str, str, Any]] = [
    ("input_directory", "Directory for input time-series files", "input"),
    (
        "input_format",
        "Preferred input file format: 'parquet' or 'csv'",
        "parquet",
    ),
    ("demand_fail_cost", "Penalty for unserved load [$/MWh]", 1000),
    ("reserve_fail_cost", "Penalty for unserved spinning reserve [$/MW]", 5000),
    ("use_line_losses", "Enable line loss modelling (true/false)", True),
    (
        "loss_segments",
        "Number of piecewise-linear loss segments (1=linear only)",
        1,
    ),
    ("use_kirchhoff", "Apply DC Kirchhoff OPF constraints (true/false)", True),
    (
        "use_single_bus",
        "Copper-plate (no network) mode – ignores all line limits (true/false)",
        False,
    ),
    (
        "kirchhoff_threshold",
        "Minimum bus voltage [kV] for Kirchhoff constraints",
        None,
    ),
    (
        "scale_objective",
        "Divide objective coefficients by this value for solver numerics",
        1000,
    ),
    (
        "scale_theta",
        "Angle variable scaling factor (default: 1000)",
        1000,
    ),
    (
        "annual_discount_rate",
        "Annual discount rate for CAPEX [p.u.] (e.g. 0.10 = 10 %)",
        0.1,
    ),
    ("output_directory", "Directory for solution output files", "output"),
    ("output_format", "Output file format: 'parquet' or 'csv'", "parquet"),
    (
        "output_compression",
        "Parquet compression codec: 'gzip', 'snappy', 'zstd', or ''",
        "gzip",
    ),
    (
        "use_lp_names",
        "Write variable/constraint names to the LP file (true/false)",
        None,
    ),
    (
        "use_uid_fname",
        "Use uid-based filenames for output (true/false)",
        None,
    ),
    ("lp_algorithm", "LP solver algorithm code (0=auto)", None),
    ("lp_threads", "Number of LP solver threads (0=auto)", None),
    ("lp_presolve", "Enable LP presolve (true/false)", None),
    ("log_directory", "Directory for solver log files", "logs"),
    # SDDP options
    (
        "sddp_solver_type",
        "SDDP sub-problem solver type: 'monolithic' or 'sddp'",
        None,
    ),
    (
        "sddp_cut_sharing_mode",
        "How Benders cuts are shared: 'none', 'expected', or 'max'",
        None,
    ),
    (
        "sddp_cut_directory",
        "Directory for SDDP Benders cut files",
        "cuts",
    ),
    ("sddp_api_enabled", "Write SDDP status JSON for monitoring (true/false)", None),
    (
        "sddp_efficiency_update_skip",
        "SDDP iterations between reservoir efficiency updates",
        None,
    ),
    ("sddp_max_iterations", "Maximum SDDP outer iterations", None),
    (
        "sddp_convergence_tol",
        "SDDP convergence tolerance (gap between bounds)",
        None,
    ),
    ("sddp_elastic_penalty", "Penalty for elastic constraint relaxation", None),
    ("sddp_alpha_min", "Minimum alpha (future cost) lower bound", None),
    ("sddp_alpha_max", "Maximum alpha (future cost) upper bound", None),
    ("sddp_cuts_input_file", "Path to pre-computed Benders cuts file", None),
    ("sddp_sentinel_file", "Path to sentinel file that stops SDDP early", None),
    (
        "sddp_elastic_mode",
        "Elastic constraint mode: 'none', 'feasibility', or 'cost'",
        None,
    ),
]

# ------------------------------------------------------------------
# Ordered sheet list (determines workbook tab order)
# ------------------------------------------------------------------
_TEMPLATE_SIMULATION_SHEETS = [
    "block_array",
    "stage_array",
    "scenario_array",
    "phase_array",
    "scene_array",
]

_TEMPLATE_SYSTEM_SHEETS = [
    "bus_array",
    "generator_array",
    "generator_profile_array",
    "demand_array",
    "demand_profile_array",
    "line_array",
    "battery_array",
    "converter_array",
    "reserve_zone_array",
    "reserve_provision_array",
    "junction_array",
    "waterway_array",
    "flow_array",
    "reservoir_array",
    "filtration_array",
    "turbine_array",
    "reservoir_efficiency_array",
]

# ------------------------------------------------------------------
# Introduction sheet content
# ------------------------------------------------------------------
_INTRO_LINES = [
    ("gtopt Planning Template", "TITLE"),
    ("", ""),
    (
        "This workbook is a ready-to-use template for building gtopt planning "
        "cases with the igtopt converter.",
        "BODY",
    ),
    (
        "Fill in each sheet with your system data, then run:  igtopt <this_file>.xlsx",
        "BODY",
    ),
    ("", ""),
    ("Sheet Guide", "HEADING"),
    ("", ""),
    ("Sheet name", "Header", "Description", "TABLE_HEADER"),
    ("options", "", "Global solver and I/O options (one row per option key)", "TABLE"),
    (
        "block_array",
        "Simulation",
        "Operating time blocks (hours)",
        "TABLE",
    ),
    (
        "stage_array",
        "Simulation",
        "Planning stages (groups of blocks)",
        "TABLE",
    ),
    (
        "scenario_array",
        "Simulation",
        "Probability-weighted scenarios",
        "TABLE",
    ),
    (
        "phase_array",
        "Simulation",
        "SDDP phases (groups of stages) — leave empty for monolithic",
        "TABLE",
    ),
    (
        "scene_array",
        "Simulation",
        "SDDP scenes (groups of scenarios) — leave empty for monolithic",
        "TABLE",
    ),
    ("bus_array", "System", "Electrical busbars (nodes)", "TABLE"),
    (
        "generator_array",
        "System",
        "Thermal / renewable / hydro generators",
        "TABLE",
    ),
    (
        "generator_profile_array",
        "System",
        "Time-varying capacity-factor profiles for generators",
        "TABLE",
    ),
    ("demand_array", "System", "Electricity consumers (loads)", "TABLE"),
    (
        "demand_profile_array",
        "System",
        "Time-varying load-shape profiles for demands",
        "TABLE",
    ),
    ("line_array", "System", "Transmission lines / branches", "TABLE"),
    (
        "battery_array",
        "System",
        "Energy storage devices (batteries, pumped-hydro, etc.)",
        "TABLE",
    ),
    (
        "converter_array",
        "System",
        "Battery charge/discharge converter links",
        "TABLE",
    ),
    ("reserve_zone_array", "System", "Spinning-reserve zones", "TABLE"),
    (
        "reserve_provision_array",
        "System",
        "Generator contributions to reserve zones",
        "TABLE",
    ),
    ("junction_array", "System", "Hydraulic junctions (nodes)", "TABLE"),
    (
        "waterway_array",
        "System",
        "Water channels between junctions",
        "TABLE",
    ),
    (
        "flow_array",
        "System",
        "Exogenous inflows / outflows at junctions",
        "TABLE",
    ),
    (
        "reservoir_array",
        "System",
        "Water reservoirs (lakes, dams)",
        "TABLE",
    ),
    (
        "filtration_array",
        "System",
        "Water seepage from waterways into reservoirs",
        "TABLE",
    ),
    (
        "turbine_array",
        "System",
        "Hydro turbines linking waterways to generators",
        "TABLE",
    ),
    (
        "reservoir_efficiency_array",
        "System",
        "Volume-dependent turbine productivity curves",
        "TABLE",
    ),
    ("", ""),
    ("Time-series Sheets (@-sheets)", "HEADING"),
    ("", ""),
    (
        "Sheets named  Element@field  (e.g. 'Demand@lmax', 'GeneratorProfile@profile')",
        "BODY",
    ),
    (
        "are written as Parquet/CSV files to the input directory. Column headers",
        "BODY",
    ),
    (
        "must be: scenario, stage, block, then one column per element (by name).",
        "BODY",
    ),
    ("", ""),
    ("Conventions", "HEADING"),
    ("", ""),
    ("• Sheets / columns whose name starts with '.' are silently skipped.", "BODY"),
    (
        "• Leave cells empty (or use NaN) for optional fields — they are omitted from JSON.",
        "BODY",
    ),
    (
        "• Fields of type  number|array|filename  accept a scalar, an inline array",
        "BODY",
    ),
    ("  (JSON syntax, e.g. [100,90,80]), or a filename stem.", "BODY"),
    (
        "• The 'active' column accepts 1 (active) or 0 (inactive); "
        "leave blank to use the default (1).",
        "BODY",
    ),
    ("", ""),
    (
        "Generated by igtopt --make-template — re-run to refresh after C++ changes.",
        "FOOTER",
    ),
]

# ------------------------------------------------------------------
# Example rows for common sheets
# ------------------------------------------------------------------
_EXAMPLES: dict[str, list[dict[str, Any]]] = {
    "block_array": [{"uid": 1, "name": "b1", "duration": 1.0}],
    "stage_array": [{"uid": 1, "name": "s1", "first_block": 0, "count_block": 8760}],
    "scenario_array": [
        {"uid": 1, "name": "sc1", "active": 1, "probability_factor": 1.0}
    ],
    "bus_array": [
        {"uid": 1, "name": "b1", "voltage": 220.0},
        {"uid": 2, "name": "b2", "voltage": 220.0},
    ],
    "generator_array": [
        {"uid": 1, "name": "g1", "bus": "b1", "pmax": 100.0, "gcost": 20.0},
        {"uid": 2, "name": "g2", "bus": "b2", "pmax": 200.0, "gcost": 35.0},
    ],
    "demand_array": [
        {"uid": 1, "name": "d1", "bus": "b1", "lmax": 80.0},
        {"uid": 2, "name": "d2", "bus": "b2", "lmax": 150.0},
    ],
    "line_array": [
        {
            "uid": 1,
            "name": "l1_2",
            "bus_a": "b1",
            "bus_b": "b2",
            "tmax_ab": 200.0,
            "tmax_ba": 200.0,
            "reactance": 0.05,
        }
    ],
    "options": [],  # handled separately — no example rows in data sheet
}


def _find_repo_root(start: pathlib.Path) -> pathlib.Path:
    """Walk up the directory tree to find the repository root."""
    for parent in [start.resolve(), *start.resolve().parents]:
        if (parent / "include" / "gtopt").is_dir():
            return parent
    return start.resolve()


# Regex to extract json_member field names from JSON headers
_FIELD_NAME_RE = re.compile(r'json_\w+<"([^"]+)"')


def parse_json_header_fields(header_path: pathlib.Path) -> list[str]:
    """Extract JSON field names from a json_*.hpp header file.

    Returns the ordered list of field names as they appear in the
    json_member_list for the *last* struct defined in the file
    (which is typically the main user-facing struct, not the Attrs variant).
    """
    text = header_path.read_text(encoding="utf-8", errors="ignore")
    fields = _FIELD_NAME_RE.findall(text)
    seen: set[str] = set()
    result = []
    for f in fields:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result


def parse_system_arrays(json_system_hpp: pathlib.Path) -> list[str]:
    """Parse json_system.hpp to get the ordered list of array field names."""
    text = json_system_hpp.read_text(encoding="utf-8", errors="ignore")
    return re.findall(r'json_array_null<"([^"]+)"', text)


def parse_simulation_arrays(json_simulation_hpp: pathlib.Path) -> list[str]:
    """Parse json_simulation.hpp to get the ordered list of array field names."""
    text = json_simulation_hpp.read_text(encoding="utf-8", errors="ignore")
    return re.findall(r'json_array_null<"([^"]+)"', text)


def _build_workbook(  # noqa: PLR0912,PLR0915
    output_path: pathlib.Path, header_dir: pathlib.Path
) -> None:
    """Build the Excel template workbook and write to *output_path*."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment,
            Font,
            PatternFill,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        print(
            f"Error: openpyxl is required.  Install it with: pip install openpyxl\n{exc}",
            file=sys.stderr,
        )
        sys.exit(1)

    # Styles
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F3864")
    HEADING_FONT = Font(name="Calibri", size=12, bold=True, color="2F5496")
    BODY_FONT = Font(name="Calibri", size=10)
    FOOTER_FONT = Font(name="Calibri", size=9, italic=True, color="808080")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")
    HELP_FILL = PatternFill("solid", fgColor="EBF3FB")
    REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
    HEADER_FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF")
    TABLE_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    TABLE_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    default_sheet.title = ".introduction"

    # ------------------------------------------------------------------
    # .introduction sheet
    # ------------------------------------------------------------------
    ws = default_sheet
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 58

    row = 1
    for entry in _INTRO_LINES:
        if entry[-1] == "TITLE":
            cell = ws.cell(row=row, column=1, value=entry[0])
            cell.font = TITLE_FONT
            ws.merge_cells(f"A{row}:C{row}")
        elif entry[-1] == "HEADING":
            cell = ws.cell(row=row, column=1, value=entry[0])
            cell.font = HEADING_FONT
            ws.merge_cells(f"A{row}:C{row}")
        elif entry[-1] == "BODY":
            cell = ws.cell(row=row, column=1, value=entry[0])
            cell.font = BODY_FONT
            ws.merge_cells(f"A{row}:C{row}")
        elif entry[-1] == "FOOTER":
            cell = ws.cell(row=row, column=1, value=entry[0])
            cell.font = FOOTER_FONT
            ws.merge_cells(f"A{row}:C{row}")
        elif entry[-1] == "TABLE_HEADER":
            for col_idx, val in enumerate(entry[:3], start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = TABLE_HEADER_FONT
                c.fill = TABLE_HEADER_FILL
        elif entry[-1] == "TABLE":
            for col_idx, val in enumerate(entry[:3], start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = BODY_FONT
                if row % 2 == 0:
                    c.fill = ALT_ROW_FILL
        elif entry[0] == "":
            pass  # blank spacer row
        row += 1

    # ------------------------------------------------------------------
    # options sheet
    # ------------------------------------------------------------------
    ws_opts = wb.create_sheet("options")
    ws_opts.column_dimensions["A"].width = 34
    ws_opts.column_dimensions["B"].width = 22
    ws_opts.column_dimensions["C"].width = 64

    # Header row
    for col_idx, header in enumerate(["option", "value", "description"], start=1):
        c = ws_opts.cell(row=1, column=col_idx, value=header)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left")

    for row_idx, (key, desc, default) in enumerate(_OPTIONS_FIELDS, start=2):
        ws_opts.cell(row=row_idx, column=1, value=key).font = Font(bold=False)
        if default is not None:
            ws_opts.cell(row=row_idx, column=2, value=default)
        ws_opts.cell(row=row_idx, column=3, value=desc).font = Font(
            italic=True, color="404040"
        )

    ws_opts.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Helper to build a data sheet
    # ------------------------------------------------------------------
    def _add_data_sheet(
        sheet_name: str,
        fields: list[tuple[str, str, bool, str, Any]],
        examples: list[dict[str, Any]] | None = None,
    ) -> None:
        ws_data = wb.create_sheet(sheet_name)

        num_fields = len(fields)
        for col_idx, (fname, ftype, required, desc, _example) in enumerate(
            fields, start=1
        ):
            # Header row
            header_cell = ws_data.cell(row=1, column=col_idx, value=fname)
            header_cell.font = HEADER_FONT_WHITE
            if required:
                header_cell.fill = HEADER_FILL  # dark blue = required
            else:
                header_cell.fill = PatternFill("solid", fgColor="2F5496")  # medium
            header_cell.alignment = Alignment(horizontal="left")

            # Help row
            help_text = f"[{ftype}]{'*' if required else ''} {desc}"
            help_cell = ws_data.cell(row=2, column=col_idx, value=help_text)
            help_cell.font = Font(name="Calibri", size=8, italic=True, color="404040")
            if required:
                help_cell.fill = REQUIRED_FILL
            else:
                help_cell.fill = HELP_FILL

        # Column widths
        for col_idx, (fname, _ft, _req, _desc, _ex) in enumerate(fields, start=1):
            col_letter = get_column_letter(col_idx)
            ws_data.column_dimensions[col_letter].width = max(14, len(fname) + 4)

        # Example rows
        if examples:
            for row_offset, example_row in enumerate(examples):
                data_row = 3 + row_offset
                for col_idx, (fname, _ft, _req, _desc, _default_ex) in enumerate(
                    fields, start=1
                ):
                    val = example_row.get(fname, None)
                    if val is not None:
                        ws_data.cell(row=data_row, column=col_idx, value=val)
        else:
            # Pre-fill a single blank example row using per-field defaults
            for col_idx, (_fname, _ft, _req, _desc, example) in enumerate(
                fields, start=1
            ):
                if example is not None:
                    ws_data.cell(row=3, column=col_idx, value=example)

        ws_data.freeze_panes = "A3"

        # Legend in a far-right column
        legend_col = num_fields + 2
        leg_label = ws_data.cell(row=1, column=legend_col, value="Legend")
        leg_label.font = Font(bold=True)
        ws_data.cell(row=2, column=legend_col, value="Dark blue = required")
        ws_data.cell(row=3, column=legend_col, value="Medium blue = optional")
        ws_data.cell(row=4, column=legend_col, value="* = required field")
        ws_data.cell(
            row=5,
            column=legend_col,
            value="number|array|filename: scalar, inline array, or filename",
        )
        ws_data.column_dimensions[get_column_letter(legend_col)].width = 46

    # ------------------------------------------------------------------
    # Simulation sheets
    # ------------------------------------------------------------------
    for sheet in _TEMPLATE_SIMULATION_SHEETS:
        fields = FIELD_META.get(sheet, [])
        if not fields:
            continue
        examples = _EXAMPLES.get(sheet)
        _add_data_sheet(sheet, fields, examples)

    # ------------------------------------------------------------------
    # System sheets
    # ------------------------------------------------------------------
    for sheet in _TEMPLATE_SYSTEM_SHEETS:
        fields = FIELD_META.get(sheet, [])
        if not fields:
            continue
        examples = _EXAMPLES.get(sheet)
        _add_data_sheet(sheet, fields, examples)

    # ------------------------------------------------------------------
    # Example time-series sheets
    # ------------------------------------------------------------------
    # Demand@lmax
    ws_ts = wb.create_sheet("Demand@lmax")
    for col_idx, header in enumerate(
        ["scenario", "stage", "block", "d1", "d2"], start=1
    ):
        c = ws_ts.cell(row=1, column=col_idx, value=header)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
    ws_ts.cell(row=2, column=1, value=1)
    ws_ts.cell(row=2, column=2, value=1)
    ws_ts.cell(row=2, column=3, value=1)
    ws_ts.cell(row=2, column=4, value=80.0)
    ws_ts.cell(row=2, column=5, value=150.0)
    ws_ts.freeze_panes = "A2"

    # GeneratorProfile@profile
    ws_ts2 = wb.create_sheet("GeneratorProfile@profile")
    for col_idx, header in enumerate(
        ["scenario", "stage", "block", "g_solar"], start=1
    ):
        c = ws_ts2.cell(row=1, column=col_idx, value=header)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
    ws_ts2.cell(row=2, column=1, value=1)
    ws_ts2.cell(row=2, column=2, value=1)
    ws_ts2.cell(row=2, column=3, value=1)
    ws_ts2.cell(row=2, column=4, value=0.75)
    ws_ts2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Template written to: {output_path}", file=sys.stderr)


def _list_sheets(header_dir: pathlib.Path) -> None:
    """Print the sheets that would be generated (does not write a file)."""
    json_dir = header_dir / "json"
    if not json_dir.is_dir():
        json_dir = header_dir

    sim_file = json_dir / "json_simulation.hpp"
    sys_file = json_dir / "json_system.hpp"

    print("# Simulation sheets (from json_simulation.hpp):")
    if sim_file.exists():
        for arr in parse_simulation_arrays(sim_file):
            print(f"  {arr}")
    else:
        for arr in _TEMPLATE_SIMULATION_SHEETS:
            print(f"  {arr}  (fallback — json_simulation.hpp not found)")

    print("\n# System sheets (from json_system.hpp):")
    if sys_file.exists():
        for arr in parse_system_arrays(sys_file):
            print(f"  {arr}")
    else:
        for arr in _TEMPLATE_SYSTEM_SHEETS:
            print(f"  {arr}  (fallback — json_system.hpp not found)")


def _run(args) -> int:
    t_start = time.monotonic()

    # Use local formatting variables so multiple calls are independent.
    pretty = getattr(args, "pretty", False)
    _indent = _PRETTY_INDENT if pretty else _COMPACT_INDENT
    _separators = _PRETTY_SEPARATORS if pretty else _COMPACT_SEPARATORS

    options: dict[str, Any] = {}
    options["input_directory"] = str(args.input_directory)
    options["input_format"] = args.input_format

    # Collect simulation and system arrays as ordered dicts so we can write
    # the nested {options, simulation, system} structure that gtopt expects.
    simulation: dict[str, Any] = {}
    system: dict[str, Any] = {"name": args.name}

    json_path = args.json_file.with_suffix(".json")
    filenames = args.filenames
    # Pre-initialise with zeros so log_conversion_stats always has every key.
    counts: dict[str, int] = {s: 0 for s in expected_sheets if s != "options"}
    for filename in filenames:
        filepath = pathlib.Path(filename)
        if filepath.is_dir() or not filepath.exists():
            filepath = filepath.with_suffix(".xlsx")
        if not filepath.exists():
            logging.info("skipping not existing file %s", filepath)
            continue

        xls = pd.read_excel(str(filepath), sheet_name=None, engine="openpyxl")
        for sheet_name, df in xls.items():
            if sheet_name[0] == ".":
                logging.info("skipping sheet %s", sheet_name)
                continue

            if "@" in sheet_name:
                [cname, fname] = sheet_name.split("@")
                input_file = df_to_file(
                    df,
                    args.input_directory,
                    cname,
                    fname,
                    args.input_format,
                    args.compression,
                )
                logging.info("sheet %s saved as %s", sheet_name, input_file)

                continue

            if sheet_name in expected_sheets:
                logging.info("processing sheet %s", sheet_name)
            else:
                if not args.parse_unexpected_sheets:
                    logging.warning("skipping unexpected sheet %s", sheet_name)
                    continue

                logging.warning("processing unexpected sheet %s", sheet_name)

            if sheet_name == "options":
                options = df_to_opts(df, options)
                continue

            df_parsed = json.loads(
                df_to_str(df, args.skip_nulls, indent=_indent, separators=_separators)
            )
            counts[sheet_name] = len(df_parsed)

            if sheet_name in _SIMULATION_SHEETS:
                simulation[sheet_name] = df_parsed
            else:
                system[sheet_name] = df_parsed

    has_data = len(simulation) > 0 or len(system) > 1  # system always has "name"
    if has_data:
        planning: dict[str, Any] = {}
        planning["options"] = options
        if simulation:
            planning["simulation"] = simulation
        planning["system"] = system

        with json_path.open("w") as json_file:
            json.dump(
                planning,
                json_file,
                indent=_indent if pretty else None,
                separators=None if pretty else _separators,
            )
            json_file.write("\n")

        logging.info("gtopt input file %s was successfully generated", str(json_path))

        elapsed = time.monotonic() - t_start
        log_conversion_stats(counts, options, elapsed)

        if getattr(args, "zip", False):
            zip_path = create_zip_output(json_path, pathlib.Path(args.input_directory))
            print(f"ZIP archive created: {zip_path}")
    else:
        logging.warning(
            "no valid data was found, the file %s was not generated", str(json_path)
        )

    return 0


def main() -> None:
    """CLI entry point: parse arguments and run igtopt conversion."""
    try:
        parser = argparse.ArgumentParser(
            prog="igtopt",
            description=_DESCRIPTION,
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog=_EPILOG,
        )
        parser.add_argument(
            dest="filenames",
            nargs="*",  # changed from "+" to allow --make-template with no XLSX
            metavar="XLSX",
            help="Excel workbook(s) to convert to gtopt input data",
        )
        parser.add_argument(
            "-j",
            "--json-file",
            type=pathlib.Path,
            metavar="FILE",
            help=(
                "output JSON file (default: <first workbook stem>.json in the "
                "current directory)"
            ),
        )
        parser.add_argument(
            "-d",
            "--input-directory",
            type=pathlib.Path,
            metavar="DIR",
            help=(
                "directory for time-series data files written from '@'-sheets "
                "(default: <first workbook stem>/)"
            ),
        )
        parser.add_argument(
            "-f",
            "--input-format",
            choices=["csv", "parquet"],
            default="parquet",
            help="file format for time-series data files (default: %(default)s)",
        )
        parser.add_argument(
            "-n",
            "--name",
            metavar="NAME",
            help=(
                "system name written to the JSON output "
                "(default: <first workbook stem>)"
            ),
        )
        parser.add_argument(
            "-c",
            "--compression",
            default="gzip",
            metavar="ALG",
            help=(
                "compression algorithm for Parquet output files "
                "(default: %(default)s); pass '' to disable compression"
            ),
        )
        parser.add_argument(
            "-p",
            "--pretty",
            action=argparse.BooleanOptionalAction,
            default=False,
            help="write JSON with indented pretty-print formatting (default: compact)",
        )
        parser.add_argument(
            "-N",
            "--skip-nulls",
            action=argparse.BooleanOptionalAction,
            default=False,
            help="omit keys with null/NaN values from the JSON output (default: include)",
        )
        parser.add_argument(
            "-U",
            "--parse-unexpected-sheets",
            action=argparse.BooleanOptionalAction,
            default=False,
            help=(
                "also process sheets whose names are not in the expected list "
                "(default: warn and skip)"
            ),
        )
        parser.add_argument(
            "-z",
            "--zip",
            action=argparse.BooleanOptionalAction,
            default=False,
            help=(
                "bundle the JSON file and all data files into a single ZIP archive "
                "(compatible with gtopt_guisrv and gtopt_websrv)"
            ),
        )
        parser.add_argument(
            "-l",
            "--log-level",
            default="INFO",
            choices=_LOG_LEVEL_CHOICES,
            metavar="LEVEL",
            help=(
                "logging verbosity: DEBUG, INFO, WARNING, ERROR, CRITICAL "
                "(default: %(default)s)"
            ),
        )
        parser.add_argument(
            "-V",
            "--version",
            action="version",
            version=f"%(prog)s {__version__}",
        )
        # Template generation options
        parser.add_argument(
            "-T",
            "--make-template",
            action="store_true",
            default=False,
            help=(
                "generate the gtopt igtopt Excel template from C++ JSON headers "
                "instead of converting a workbook; see also --header-dir, --list-sheets"
            ),
        )
        parser.add_argument(
            "--header-dir",
            metavar="DIR",
            help=(
                "path to include/gtopt/ (auto-detected from the repo root if omitted); "
                "only used with --make-template"
            ),
        )
        parser.add_argument(
            "--list-sheets",
            action="store_true",
            default=False,
            help=(
                "print the list of sheets derived from C++ headers and exit; "
                "only used with --make-template"
            ),
        )

        args = parser.parse_args()

        logging.basicConfig(
            level=getattr(logging, args.log_level),
            format="%(asctime)s %(levelname)s %(message)s",
        )

        # --make-template mode: generate the Excel template and exit
        if args.make_template or args.list_sheets:
            repo_root = _find_repo_root(pathlib.Path(__file__).parent)
            if args.header_dir:
                header_dir = pathlib.Path(args.header_dir)
            else:
                header_dir = repo_root / "include" / "gtopt"

            if not header_dir.is_dir():
                logging.error(
                    "header directory '%s' not found; "
                    "use --header-dir to specify the path to include/gtopt/",
                    header_dir,
                )
                sys.exit(1)

            if args.list_sheets:
                _list_sheets(header_dir)
                return

            if args.json_file:
                output_path = pathlib.Path(args.json_file)
            else:
                output_path = repo_root / "docs" / "templates" / "gtopt_template.xlsx"

            output_path.parent.mkdir(parents=True, exist_ok=True)
            _build_workbook(output_path, header_dir)
            logging.info("Template written to %s", output_path)
            return

        # Normal conversion mode: require at least one XLSX file
        if not args.filenames:
            parser.error("at least one XLSX file is required (or use --make-template)")

        if not args.json_file:
            args.json_file = pathlib.Path(args.filenames[0]).with_suffix(".json")
            logging.info("using json_file %s", args.json_file)

        if not args.input_directory:
            args.input_directory = pathlib.Path(args.filenames[0]).stem
            logging.info("using input_directory %s", args.input_directory)

        if not args.name:
            args.name = pathlib.Path(args.filenames[0]).stem
            logging.info("using system name %s", args.name)

        result = _run(args)
    except (IOError, ValueError, argparse.ArgumentError) as e:
        logging.error("unexpected error: %s", str(e))
        sys.exit(1)

    sys.exit(result)


if __name__ == "__main__":
    main()
