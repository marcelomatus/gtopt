# SPDX-License-Identifier: BSD-3-Clause
"""Tests for gtopt_timeseries_export."""

from __future__ import annotations


import openpyxl
import pytest

from gtopt_timeseries_export import export_from_dict, export_to_excel
from gtopt_timeseries_export.main import main as cli_main


def _sample_results() -> dict:
    return {
        "solution": {"obj_value": "2.5", "status": "0"},
        "outputs": {
            "Generator/generation_sol": {
                "columns": ["scenario", "stage", "block", "uid:1"],
                "data": [
                    [1, 1, 1, 10.0],
                    [1, 1, 2, 20.0],
                ],
            },
            "Demand/fail_sol": {
                "columns": ["scenario", "stage", "block", "uid:2"],
                "data": [
                    [1, 1, 1, 0.0],
                    [1, 1, 2, 1.0],
                ],
            },
        },
    }


def test_export_from_dict(tmp_path):
    out = tmp_path / "result.xlsx"
    path = export_from_dict(_sample_results(), out, scale_objective=1000.0)
    assert path == out
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert "Overview" in wb.sheetnames
    # The Overview sheet must list at least Status and Objective metrics
    overview_cells = [
        (row[0].value, row[1].value if len(row) > 1 else None)
        for row in wb["Overview"].iter_rows(min_row=2)
    ]
    metrics = {k: v for (k, v) in overview_cells if k}
    assert "Status" in metrics
    assert "Objective (scaled)" in metrics
    assert "Total generation (MWh)" in metrics
    assert metrics["Objective (scaled)"] == pytest.approx(2500.0)

    # Per-output sheets are present
    sheet_names = [s.lower() for s in wb.sheetnames]
    assert any("generator" in s for s in sheet_names)
    assert any("demand" in s for s in sheet_names)


def test_export_to_excel_dir(tmp_path):
    # Create a minimal output directory
    outdir = tmp_path / "out"
    outdir.mkdir()
    (outdir / "solution.csv").write_text("obj_value,1.5\nstatus,0\n")
    gen_dir = outdir / "Generator"
    gen_dir.mkdir()
    (gen_dir / "generation_sol.csv").write_text(
        "scenario,stage,block,uid:1\n1,1,1,5.0\n"
    )
    dest = tmp_path / "wb.xlsx"
    result = export_to_excel(outdir, dest)
    assert result == dest
    assert dest.exists()


def test_export_nonexistent_source(tmp_path):
    with pytest.raises(FileNotFoundError):
        export_to_excel(tmp_path / "missing", tmp_path / "out.xlsx")


def test_export_wrong_suffix(tmp_path):
    fake = tmp_path / "data.txt"
    fake.write_text("hi")
    with pytest.raises(ValueError):
        export_to_excel(fake, tmp_path / "out.xlsx")


def test_cli(tmp_path, capsys):
    outdir = tmp_path / "out"
    outdir.mkdir()
    (outdir / "solution.csv").write_text("obj_value,1.0\nstatus,0\n")
    dest = tmp_path / "cli.xlsx"
    code = cli_main([str(outdir), "-o", str(dest)])
    assert code == 0
    assert dest.exists()
    captured = capsys.readouterr()
    assert str(dest) in captured.out


def test_sheet_name_sanitization(tmp_path):
    results = {
        "solution": {},
        "outputs": {
            "a/b:c?d*e[f]g\\h/i": {
                "columns": ["col"],
                "data": [[1]],
            },
        },
    }
    dest = tmp_path / "safe.xlsx"
    export_from_dict(results, dest)
    wb = openpyxl.load_workbook(dest)
    # None of the sheet names may contain Excel-invalid characters
    for name in wb.sheetnames:
        for bad in r":\/?*[]":
            assert bad not in name


# ---------------------------------------------------------------------------
# Additional coverage: Overview sheet content
# ---------------------------------------------------------------------------


def test_overview_metrics_from_dict(tmp_path):
    """Overview sheet must contain Total unserved (MWh) and Total generation rows."""
    results = {
        "solution": {"obj_value": "1.0", "status": "0"},
        "outputs": {
            "Generator/generation_sol": {
                "columns": ["scenario", "stage", "block", "uid:1"],
                "data": [[1, 1, 1, 50.0]],
            },
            "Demand/fail_sol": {
                "columns": ["scenario", "stage", "block", "uid:1"],
                "data": [[1, 1, 1, 2.5]],
            },
        },
    }
    dest = tmp_path / "out.xlsx"
    export_from_dict(results, dest, scale_objective=1.0)
    wb = openpyxl.load_workbook(dest)
    rows_by_key = {
        row[0].value: row[1].value if len(row) > 1 else None
        for row in wb["Overview"].iter_rows(min_row=2)
        if row[0].value
    }
    # The exporter uses "Total unserved (MWh)" and "Total generation (MWh)"
    assert "Total unserved (MWh)" in rows_by_key
    assert rows_by_key["Total unserved (MWh)"] == pytest.approx(2.5)
    assert rows_by_key["Total generation (MWh)"] == pytest.approx(50.0)


def test_multiple_output_sheets(tmp_path):
    """Every entry in outputs should produce a sheet."""
    results = {
        "solution": {},
        "outputs": {
            "Generator/generation_sol": {
                "columns": ["s", "t", "b", "uid:1"],
                "data": [[1, 1, 1, 10.0]],
            },
            "Demand/load_sol": {
                "columns": ["s", "t", "b", "uid:2"],
                "data": [[1, 1, 1, 8.0]],
            },
            "Bus/balance_dual": {
                "columns": ["s", "t", "b", "uid:3"],
                "data": [[1, 1, 1, 20.0]],
            },
        },
    }
    dest = tmp_path / "multi.xlsx"
    export_from_dict(results, dest)
    wb = openpyxl.load_workbook(dest)
    # One sheet per output table + Overview
    assert len(wb.sheetnames) >= 4


def test_column_headers_match_csv_columns(tmp_path):
    """Column headers in each data sheet must match the columns field."""
    results = {
        "solution": {},
        "outputs": {
            "Generator/generation_sol": {
                "columns": ["scenario", "stage", "block", "uid:42"],
                "data": [[1, 1, 1, 77.0]],
            }
        },
    }
    dest = tmp_path / "headers.xlsx"
    export_from_dict(results, dest)
    wb = openpyxl.load_workbook(dest)
    # Find the generator sheet (sheet name derived from key)
    gen_sheet = next(
        (s for s in wb.sheetnames if "Generator" in s or "generation" in s.lower()),
        None,
    )
    assert gen_sheet is not None
    ws = wb[gen_sheet]
    headers = [cell.value for cell in ws[1]]
    assert "uid:42" in headers
    assert "scenario" in headers


def test_autofit_does_not_crash_empty_sheet(tmp_path):
    """Export with a table that has zero data rows should not crash."""
    results = {
        "solution": {},
        "outputs": {
            "Generator/generation_sol": {
                "columns": ["scenario", "stage", "block"],
                "data": [],
            }
        },
    }
    dest = tmp_path / "empty.xlsx"
    export_from_dict(results, dest)
    assert dest.exists()


def test_scale_objective_reflected_in_overview(tmp_path):
    results = {
        "solution": {"obj_value": "3.0", "status": "0"},
        "outputs": {},
    }
    dest = tmp_path / "scaled.xlsx"
    export_from_dict(results, dest, scale_objective=500.0)
    wb = openpyxl.load_workbook(dest)
    rows = {
        row[0].value: row[1].value
        for row in wb["Overview"].iter_rows(min_row=2)
        if row[0].value
    }
    # 3.0 * 500 = 1500
    assert rows.get("Objective (scaled)") == pytest.approx(1500.0)


def test_cli_with_scale_objective(tmp_path):
    outdir = tmp_path / "out"
    outdir.mkdir()
    (outdir / "solution.csv").write_text("obj_value,2.0\nstatus,0\n")
    dest = tmp_path / "scaled.xlsx"
    code = cli_main([str(outdir), "-o", str(dest), "--scale-objective", "1000"])
    assert code == 0
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    rows = {
        row[0].value: row[1].value
        for row in wb["Overview"].iter_rows(min_row=2)
        if row[0].value
    }
    assert rows.get("Objective (scaled)") == pytest.approx(2000.0)
