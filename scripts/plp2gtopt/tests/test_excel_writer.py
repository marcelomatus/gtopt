"""Tests for plp2gtopt Excel output (--excel-output / -E flag)."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from unittest.mock import patch

import pytest

from plp2gtopt.excel_writer import (
    _collect_field_order,
    _flatten_options,
    _records_to_df,
    build_plp_excel,
)
from plp2gtopt.main import build_options, make_parser
from plp2gtopt.plp2gtopt import convert_plp_case

# ---------------------------------------------------------------------------
# Paths to fixture cases
# ---------------------------------------------------------------------------

_CASES_DIR = Path(__file__).parent.parent.parent / "cases"
_PLPMin1Bus = _CASES_DIR / "plp_min_1bus"


def _make_opts(input_dir: Path, tmp_path: Path, case_name: str = "test") -> dict:
    out_dir = tmp_path / case_name
    out_dir.mkdir(parents=True, exist_ok=True)
    return {
        "input_dir": input_dir,
        "output_dir": out_dir,
        "output_file": out_dir / f"{case_name}.json",
        "excel_file": tmp_path / f"{case_name}.xlsx",
        "excel_output": True,
        "hydrologies": "1",
        "last_stage": -1,
        "last_time": -1,
        "compression": "zstd",
        "probability_factors": None,
        "discount_rate": 0.0,
        "management_factor": 0.0,
    }


# ---------------------------------------------------------------------------
# Unit tests for helper functions
# ---------------------------------------------------------------------------


def test_flatten_options_basic():
    """_flatten_options converts a flat dict to (key, value) pairs."""
    opts = {"input_directory": "input", "use_kirchhoff": False}
    pairs = dict(_flatten_options(opts))
    assert pairs["input_directory"] == "input"
    assert pairs["use_kirchhoff"] is False


def test_flatten_options_nested():
    """_flatten_options flattens nested dicts (like sddp_options)."""
    opts = {
        "input_directory": "input",
        "sddp_options": {"sddp_solver_type": "sddp", "sddp_max_iterations": 10},
    }
    pairs = dict(_flatten_options(opts))
    assert pairs["input_directory"] == "input"
    assert pairs["sddp_solver_type"] == "sddp"
    assert pairs["sddp_max_iterations"] == 10


def test_flatten_options_skips_none():
    """_flatten_options omits None values."""
    opts = {"input_directory": "input", "kirchhoff_threshold": None}
    pairs = dict(_flatten_options(opts))
    assert "kirchhoff_threshold" not in pairs


def test_records_to_df_preserves_columns():
    """_records_to_df converts records to a DataFrame with all keys."""
    records = [{"uid": 1, "name": "b1"}, {"uid": 2, "name": "b2"}]
    df = _records_to_df(records, "bus_array")
    assert "uid" in df.columns
    assert "name" in df.columns
    assert len(df) == 2


def test_records_to_df_empty():
    """_records_to_df returns empty DataFrame for empty records."""
    df = _records_to_df([], "bus_array")
    assert df.empty


def test_records_to_df_serializes_list_values():
    """_records_to_df serializes list values as JSON strings."""
    records = [{"uid": 1, "segments": [{"volume": 10, "slope": 1.0}]}]
    df = _records_to_df(records, "filtration_array")
    val = df.loc[0, "segments"]
    assert isinstance(val, str)

    parsed = json.loads(val)
    assert parsed[0]["volume"] == 10
    assert parsed[0]["slope"] == 1.0


def test_collect_field_order_returns_list():
    """_collect_field_order returns a list of strings for known sheet names."""
    order = _collect_field_order("bus_array")
    # If igtopt is installed, should return known fields
    assert isinstance(order, list)


# ---------------------------------------------------------------------------
# build_plp_excel unit tests
# ---------------------------------------------------------------------------

pytest.importorskip("openpyxl")


def _minimal_planning() -> dict:
    """Return a minimal planning dict for unit testing."""
    return {
        "options": {
            "input_directory": "test_input",
            "input_format": "parquet",
            "output_directory": "results",
            "output_format": "parquet",
            "use_kirchhoff": False,
            "use_single_bus": False,
            "demand_fail_cost": 1000,
            "scale_objective": 1000,
        },
        "simulation": {
            "block_array": [{"uid": 1, "name": "b1", "duration": 1.0}],
            "stage_array": [
                {"uid": 1, "name": "s1", "first_block": 0, "count_block": 1}
            ],
            "scenario_array": [{"uid": 1, "name": "sc1", "probability_factor": 1.0}],
        },
        "system": {
            "name": "test",
            "bus_array": [{"uid": 1, "name": "Bus1", "voltage": 220.0}],
            "generator_array": [
                {"uid": 1, "name": "G1", "bus": 1, "pmax": 100.0, "gcost": 50.0}
            ],
            "demand_array": [{"uid": 1, "name": "D1", "bus": 1, "lmax": "lmax"}],
        },
    }


def test_build_plp_excel_creates_file(tmp_path):
    """build_plp_excel creates a .xlsx file."""
    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {"input_dir": "/tmp/x"})
    assert output_path.exists()
    assert output_path.stat().st_size > 0


def test_build_plp_excel_has_intro_sheet(tmp_path):
    """Excel workbook contains .introduction sheet."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    assert ".introduction" in wb.sheetnames


def test_build_plp_excel_has_plpinfo_sheet(tmp_path):
    """Excel workbook contains .plpinfo sheet."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {"input_dir": "/data/plp"})
    wb = openpyxl.load_workbook(output_path)
    assert ".plpinfo" in wb.sheetnames


def test_build_plp_excel_has_options_sheet(tmp_path):
    """Excel workbook contains options sheet."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    assert "options" in wb.sheetnames


def test_build_plp_excel_has_simulation_sheets(tmp_path):
    """Excel workbook contains simulation array sheets."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    sheets = set(wb.sheetnames)
    assert "block_array" in sheets
    assert "stage_array" in sheets
    assert "scenario_array" in sheets


def test_build_plp_excel_has_system_sheets(tmp_path):
    """Excel workbook contains system array sheets."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    sheets = set(wb.sheetnames)
    assert "bus_array" in sheets
    assert "generator_array" in sheets
    assert "demand_array" in sheets


def test_build_plp_excel_data_rows(tmp_path):
    """Excel data sheets have correct header and data rows."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)

    ws = wb["bus_array"]
    # Row 1 should be the header
    headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
    assert "uid" in headers
    assert "name" in headers
    # Row 2 should be first data row
    uid_col = headers.index("uid") + 1
    assert ws.cell(row=2, column=uid_col).value == 1


def test_build_plp_excel_options_key_value(tmp_path):
    """Options sheet has option | value columns."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    ws = wb["options"]
    # Row 1 is the header
    assert ws.cell(row=1, column=1).value == "option"
    assert ws.cell(row=1, column=2).value == "value"


def test_build_plp_excel_at_sheets_from_parquet(tmp_path):
    """Excel workbook includes @-sheets for Parquet files in output_dir.

    uid:N columns are renamed to element names (e.g. 'D1') in the Excel
    output.  igtopt resolves them back to uid:N when reading the workbook.
    """
    import numpy as np
    import openpyxl
    import pandas as pd

    # Create a fake Parquet file with uid:1 column
    demand_dir = tmp_path / "Demand"
    demand_dir.mkdir()
    df = pd.DataFrame(
        {
            "stage": np.array([1, 1], dtype=np.int32),
            "block": np.array([1, 2], dtype=np.int32),
            "uid:1": [100.0, 90.0],
        }
    )
    df.to_parquet(demand_dir / "lmax.parquet", index=False)

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})

    wb = openpyxl.load_workbook(output_path)
    assert "Demand@lmax" in wb.sheetnames

    ws = wb["Demand@lmax"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert "stage" in headers
    assert "block" in headers
    # uid:1 is renamed to element name "D1" (from demand_array uid=1, name="D1")
    assert "D1" in headers, (
        f"Expected element name 'D1' (not 'uid:1') in @-sheet headers, got {headers}"
    )
    assert "uid:1" not in headers, "uid:N should be replaced by element name in Excel"


def test_build_plp_excel_skips_empty_sheets(tmp_path):
    """Excel workbook does not add empty array sheets."""
    import openpyxl

    planning = _minimal_planning()
    # phase_array is absent in minimal planning
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {})
    wb = openpyxl.load_workbook(output_path)
    # phase_array should NOT be present (it's not in the planning)
    assert "phase_array" not in wb.sheetnames


def test_build_plp_excel_plpinfo_contains_date(tmp_path):
    """The .plpinfo sheet contains a conversion date."""
    import openpyxl

    planning = _minimal_planning()
    output_path = tmp_path / "test.xlsx"
    build_plp_excel(planning, tmp_path, output_path, {"input_dir": "/tmp/plp"})
    wb = openpyxl.load_workbook(output_path)
    ws = wb[".plpinfo"]

    # Find a cell containing a date-like value
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=r, column=c).value is not None
    ]
    # conversion_date value should be present
    assert any("conversion_date" in str(v) for v in all_values)


# ---------------------------------------------------------------------------
# Integration tests: convert_plp_case with excel_output=True
# ---------------------------------------------------------------------------


def test_convert_plp_case_excel_output_creates_xlsx(tmp_path):
    """convert_plp_case with excel_output=True creates a .xlsx file."""
    pytest.importorskip("openpyxl")
    opts = _make_opts(_PLPMin1Bus, tmp_path)
    convert_plp_case(opts)
    excel_file = opts["excel_file"]
    assert excel_file.exists(), f"Excel file not created: {excel_file}"
    assert excel_file.stat().st_size > 0


def test_convert_plp_case_excel_output_no_json(tmp_path):
    """In excel_output mode, the JSON file is NOT written."""
    pytest.importorskip("openpyxl")
    opts = _make_opts(_PLPMin1Bus, tmp_path)
    convert_plp_case(opts)
    json_file = opts["output_file"]
    assert not json_file.exists(), "JSON should not be written in excel_output mode"


def test_convert_plp_case_excel_has_simulation_data(tmp_path):
    """Excel workbook from plp_min_1bus has block_array and stage_array sheets."""
    import openpyxl

    pytest.importorskip("openpyxl")
    opts = _make_opts(_PLPMin1Bus, tmp_path)
    convert_plp_case(opts)

    wb = openpyxl.load_workbook(opts["excel_file"])
    sheets = set(wb.sheetnames)
    assert "block_array" in sheets
    assert "stage_array" in sheets
    assert "scenario_array" in sheets


def test_convert_plp_case_excel_has_demand_at_sheet(tmp_path):
    """Excel workbook from plp_min_1bus has Demand@lmax @-sheet."""
    import openpyxl

    pytest.importorskip("openpyxl")
    opts = _make_opts(_PLPMin1Bus, tmp_path)
    convert_plp_case(opts)

    wb = openpyxl.load_workbook(opts["excel_file"])
    assert "Demand@lmax" in wb.sheetnames


def test_convert_plp_case_excel_has_generator_sheet(tmp_path):
    """Excel workbook from plp_min_1bus has generator_array sheet."""
    import openpyxl

    pytest.importorskip("openpyxl")
    opts = _make_opts(_PLPMin1Bus, tmp_path)
    convert_plp_case(opts)

    wb = openpyxl.load_workbook(opts["excel_file"])
    assert "generator_array" in wb.sheetnames


# ---------------------------------------------------------------------------
# CLI argument parsing tests
# ---------------------------------------------------------------------------


def test_make_parser_has_excel_output_flag():
    """make_parser() registers --excel-output / -E flag."""
    p = make_parser()
    args = p.parse_args(["-E"])
    assert args.excel_output is True


def test_make_parser_excel_output_default_false():
    """make_parser() defaults excel_output to False."""
    p = make_parser()
    args = p.parse_args([])
    assert args.excel_output is False


def test_make_parser_excel_file_flag():
    """make_parser() registers --excel-file / -x flag."""
    p = make_parser()
    args = p.parse_args(["-E", "-x", "/tmp/out.xlsx"])
    assert args.excel_output is True
    assert args.excel_file == Path("/tmp/out.xlsx")


def test_build_options_excel_output_default():
    """build_options() sets excel_output=False by default."""
    args = make_parser().parse_args([])
    opts = build_options(args)
    assert opts["excel_output"] is False


def test_build_options_excel_output_true():
    """build_options() sets excel_output=True when -E is passed."""
    args = make_parser().parse_args(["-E"])
    opts = build_options(args)
    assert opts["excel_output"] is True


def test_build_options_excel_file_default():
    """build_options() sets excel_file=None by default."""
    args = make_parser().parse_args([])
    opts = build_options(args)
    assert opts["excel_file"] is None


def test_build_options_excel_file_custom():
    """build_options() passes --excel-file to opts."""
    args = make_parser().parse_args(["-E", "-x", "mycase.xlsx"])
    opts = build_options(args)
    assert opts["excel_file"] == Path("mycase.xlsx")


def test_main_excel_output_creates_xlsx(tmp_path):
    """main() with -E creates an Excel workbook at the specified path."""
    pytest.importorskip("openpyxl")
    out_dir = tmp_path / "mycase"
    excel_path = tmp_path / "mycase.xlsx"

    test_argv = [
        "plp2gtopt",
        "-E",
        "-i",
        str(_PLPMin1Bus),
        "-o",
        str(out_dir),
        "-x",
        str(excel_path),
    ]
    with patch.object(sys, "argv", test_argv):
        from plp2gtopt.main import main

        main()

    assert excel_path.exists(), "Excel file not created by main() -E"
    assert excel_path.stat().st_size > 0


def test_main_excel_output_default_path(tmp_path):
    """main() with -E derives the .xlsx path from the output directory name."""
    pytest.importorskip("openpyxl")
    out_dir = tmp_path / "mycase"

    test_argv = [
        "plp2gtopt",
        "-E",
        "-i",
        str(_PLPMin1Bus),
        "-o",
        str(out_dir),
    ]
    with patch.object(sys, "argv", test_argv):
        from plp2gtopt.main import main

        main()

    # Default excel_file = output_file.with_suffix(".xlsx") = "mycase.xlsx"
    excel_path = tmp_path / "mycase.xlsx"
    assert excel_path.exists(), f"Default Excel path not found: {excel_path}"
